import numpy as np  # pragma: no cover
import os  # pragma: no cover
import pickle  # pragma: no cover

from metabolabpy.nmr import nmrData as nd  # pragma: no cover
from metabolabpy.nmr import nmrPreProc as npp  # pragma: no cover
import math  # pragma: no cover
from openpyxl import Workbook  # pragma: no cover
from string import ascii_uppercase  # pragma: no cover
import itertools  # pragma: no cover
import webbrowser  # pragma: no cover
import matplotlib  # pragma: no cover
import matplotlib.pyplot as pl  # pragma: no cover
from metabolabpy.nmr import nmrConfig  # pragma: no cover
import metabolabpy.__init__ as ml_version  # pragma: no cover
import shutil  # pragma: no cover
import pandas as pd  # pragma: no cover
import darkdetect  # pragma: no cover
import mat73
from copy import copy
import gc
from scipy.io import loadmat



class NmrDataSet:

    def __init__(self):
        self.__version__ = ml_version.__version__
        self.nmrdat = [[]]
        self.s = 0
        self.e = -1
        self.pp = npp.NmrPreProc()
        self.deselect = np.array([])
        self.deselect2 = np.array([])
        self.cmd_buffer = np.array([])
        self.cmd_idx = -1
        self.script = ""
        self.console = ""
        self.file_format_version = 0.2
        self.keep_zoom = True
        self.cf = nmrConfig.NmrConfig()
        self.cf.read_config()
        self.spcpl = []
        self.tmsp_conc = 0.5
        self.ver = ml_version.__version__
        self.int_all_exps = True
        self.int_all_data_sets = False
        self.quantify = True
        self.export_peak_excel = True
        self.export_peak_file = 'concentrations.xlsx'
        self.export_peak_path = os.path.expanduser('~')
        self.peak_fill = False
        self.file_version = 0.0
        self.wb = []
        self.old_data_set = -1
        self.old_data_exp = -1
        self.hsqc_spin_sys_connected = True
        self.data_set_name = ''
        self.data_set_number = ''
        self.baseline_algs = ['irsqr', 'arpls', 'asls', 'aspls', 'derpsalsa', 'drpls', 'iarpls', 'iasls', 'psalsa',
                              'mpls',
                              'mor', 'imor', 'mormol', 'amormol', 'rolling_ball', 'mwmv', 'tophat', 'mpspline', 'jbcd',
                              'noise_median', 'snip', 'adaptive_minimax', 'swima', 'ipsa', 'ria', 'dietrich',
                              'std_distribution', 'fastchrom', 'cwt_br', 'fabc', 'beads', 'poly', 'modpoly', 'imodpoly',
                              'penalized_poly', 'quant_reg', 'goldindec', ]
        self.default_baseline_alg = 'rolling_ball'
        self.default_lam = 1e5
        self.default_max_iter = 50
        self.default_alpha = 0.1
        self.default_beta = 10
        self.default_gamma = 15
        self.default_beta_mult = 0.98
        self.default_gamma_mult = 0.94
        self.default_half_window = 4096
        self.default_smooth_half_window = 16
        self.default_quantile = 0.3
        self.default_poly_order = 4
        self.default_add_ext = 2
        self.int1 = 0.0
        self.int2 = 0.0
        self.int3 = 0.0
        self.int4 = 0.0
        self.print_colours = []
        self.print_neg_colours = []
        self.background_colour = []
        self.foreground_colour = []
        self.plot_index = {}
        self.internal_std = 'TMSP'
        if self.cf.mode == 'light':
            self.background_colour = (255 / 255, 255 / 255, 255 / 255)
            self.foreground_colour = (0 / 255, 0 / 255, 0 / 255)
        else:
            self.background_colour = (42 / 255, 42 / 255, 42 / 255)
            self.foreground_colour = (255 / 255, 255 / 255, 255 / 255)

        # end __init__

    def add_peak(self, start_end=np.array([], dtype='float64'), peak_label=''):
        if len(start_end) != 2:
            return

        if self.int_all_data_sets:
            ds = range(len(self.nmrdat))
        else:
            ds = [self.s]

        for k in ds:
            if self.int_all_exps:
                exps = range(len(self.nmrdat[k]))
            else:
                exps = [self.e]

            for l in exps:
                self.nmrdat[k][l].add_peak(start_end, peak_label)

        # end add_peak

    def add_tmsp(self, m0=1, r2=1):
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].add_tmsp(m0=m0, r2=r2)

    # end add_tmsp

    def autophase1d_bl(self, upper_min=10.0, lower_max=-0.5):
        ref_spc_no = self.nmrdat[0][0].display.ph_ref_exp - 1
        if ref_spc_no == -1:
            return

        if self.e != ref_spc_no:
            self.nmrdat[self.s][ref_spc_no].proc_spc1d()
            ref_spc = np.copy(self.nmrdat[self.s][ref_spc_no].spc[0])
            self.nmrdat[self.s][self.e].proc.ph0[0] = self.nmrdat[self.s][ref_spc_no].proc.ph0[0]
            self.nmrdat[self.s][self.e].proc.ph1[0] = self.nmrdat[self.s][ref_spc_no].proc.ph1[0]
            self.nmrdat[self.s][self.e].proc_spc1d()
            self.nmrdat[self.s][self.e].auto_ref()
            self.nmrdat[self.s][self.e].autophase1d_bl(upper_min, lower_max, ref_spc)
            self.nmrdat[self.s][self.e].auto_ref()


    def autophase1d_bl_all(self, upper_min=10.0, lower_max=-0.5):
        ref_spc_no = self.nmrdat[self.s][0].display.ph_ref_exp - 1
        if ref_spc_no == -1:
            return

        self.nmrdat[self.s][ref_spc_no].proc_spc1d()
        ref_spc = np.copy(self.nmrdat[self.s][ref_spc_no].spc[0])
        for k in range(len(self.nmrdat[self.s])):
            if k != ref_spc_no:
                self.nmrdat[self.s][k].proc.ph0[0] = self.nmrdat[self.s][ref_spc_no].proc.ph0[0]
                self.nmrdat[self.s][k].proc.ph1[0] = self.nmrdat[self.s][ref_spc_no].proc.ph1[0]
                self.nmrdat[self.s][k].auto_ref()
                self.nmrdat[self.s][k].proc_spc1d()
                self.nmrdat[self.s][k].autophase1d_bl(upper_min, lower_max, ref_spc)
                self.nmrdat[self.s][k].auto_ref()

    def set_peak(self, start_peak, end_peak, peak_label, n_protons):
        if self.int_all_data_sets:
            ds = range(len(self.nmrdat))
        else:
            ds = [self.s]

        for k in ds:
            if self.int_all_exps:
                exps = range(len(self.nmrdat[k]))
            else:
                exps = [self.e]

            for l in exps:
                self.nmrdat[k][l].set_peak(start_peak, end_peak, peak_label, n_protons)

        # end add_peak

    def clear_peak(self):
        if self.int_all_data_sets:
            ds = range(len(self.nmrdat))
        else:
            ds = [self.s]

        for k in ds:
            if self.int_all_exps:
                exps = range(len(self.nmrdat[k]))
            else:
                exps = [self.e]

            for l in exps:
                self.nmrdat[k][l].clear_peak()

        # end add_peak

    def __str__(self):  # pragma: no cover
        r_string = '______________________________________________________________________________________\n'
        if len(self.nmrdat[self.s]) > 0:
            r_string += '\nMetaboLabPy NMR Data Set (ver. ' + self.nmrdat[self.s][self.e].ver + ')\n'
        else:
            r_string += '\nMetaboLabPy NMR Data Set (empty data set)\n'

        r_string += '______________________________________________________________________________________\n'
        r_string += '\n Number of data sets\t\t\t\t: {:0.0f}\n'.format(len(self.nmrdat))
        r_string += ' Number of NMR spectra in current data set\t: {:0.0f}\n'.format(len(self.nmrdat[self.s]))
        if (len(self.nmrdat[self.s]) > 0):
            r_string += ' Current data set/exp\t\t\t\t: {:0.0f}/{:0.0f}\n'.format(self.s + 1, self.e + 1)
            r_string += '______________________________________________________________________________________\n'
            r_string += '\nCurrent title file: \n'
            r_string += self.nmrdat[self.s][self.e].title
        return r_string
        # end __str__

    def adaptive_lb(self):
        lw = []
        for k in range(len(self.nmrdat[self.s])):
            lw.append(self.nmrdat[self.s][k].tmsp_linewidth)

        if len(np.where(np.array(lw) == 0.0)[0]) > 0:
            self.fit_tmsp_all()

        lw = []
        for k in range(len(self.nmrdat[self.s])):
            lw.append(self.nmrdat[self.s][k].tmsp_linewidth)

        lw = np.array(lw)
        print(f'lw: {lw}')
        max_idx = np.where(lw == np.max(lw))[0][0]
        max_lw = lw[max_idx]
        max_lb = self.nmrdat[self.s][max_idx].proc.lb[0]
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].proc.lb[0] = max_lb
            tmsp_lw = self.nmrdat[self.s][k].tmsp_linewidth
            if tmsp_lw != max_lw:
                self.nmrdat[self.s][k].proc.lb[0] = max_lb * 1.0 * max_lw / tmsp_lw
                self.nmrdat[self.s][k].proc_spc1d()

    # end adaptive_lb

    def autobaseline1d(self, alg='rolling_ball', lam=1000000, max_iter=50, alpha=0.1, beta=10, gamma=15, beta_mult=0.98,
                       gamma_mult=0.94, half_window=4096, quantile=0.3, poly_order=4, smooth_half_window=16, add_ext=2):
        if len(self.nmrdat) > 0:
            if len(self.nmrdat[self.s]) > 0:
                self.nmrdat[self.s][self.e].autobaseline1d(lam=lam, alg=alg, max_iter=max_iter, alpha=alpha, beta=beta,
                                                           gamma=gamma, beta_mult=beta_mult, gamma_mult=gamma_mult,
                                                           half_window=half_window, quantile=quantile,
                                                           poly_order=poly_order, smooth_half_window=smooth_half_window,
                                                           add_ext=add_ext)

        # end autobaseline1d

    def autobaseline1d_all(self):
        n_exp = len(self.nmrdat[self.s])
        orig_exp = self.e
        for k in range(n_exp):
            self.e = k
            self.autobaseline1d()

        self.e = orig_exp
        return "Finished autobaseline1d_all"
        # end autobaseline1d_all

    def autophase1d(self, width=128, num_windows=1024, max_peaks=1000, noise_fact=20):
        if len(self.nmrdat) > 0:
            if len(self.nmrdat[self.s]) > 0:
                self.nmrdat[self.s][self.e].autophase1d(width, num_windows, max_peaks, noise_fact)

        # end autophase1d

    def autophase1d1(self, gamma_factor=1.0):
        if len(self.nmrdat) > 0:
            if len(self.nmrdat[self.s]) > 0:
                self.nmrdat[self.s][self.e].autophase1d1(gamma_factor=gamma_factor)

        # end autophase1d

    def autophase1d_all(self, width=128, num_windows=1024, max_peaks=1000, noise_fact=20):
        n_exp = len(self.nmrdat[self.s])
        orig_exp = self.e
        for k in range(n_exp):
            self.e = k
            self.autophase1d(width, num_windows, max_peaks, noise_fact)

        self.e = orig_exp
        return "Finished autophase1d_all"
        # end autophase1d_all

    def autophase1d1_all(self, gamma_factor=1.0):
        n_exp = len(self.nmrdat[self.s])
        orig_exp = self.e
        for k in range(n_exp):
            self.e = k
            self.autophase1d1(gamma_factor=gamma_factor)

        self.e = orig_exp
        return "Finished autophase1d1_all"
        # end autophase1d_all

    def autophase1d_exclude_water(self, delta_sw=-1):
        n_exps = len(self.nmrdat[self.s])
        for k in range(n_exps):
            self.nmrdat[self.s][k].autophase1d_exclude_water(delta_sw)

        # end autophase1d_exclude_water

    def autophase1d_include_water(self):
        n_exps = len(self.nmrdat[self.s])
        for k in range(n_exps):
            self.nmrdat[self.s][k].autophase1d_include_water()

        # end autophase1d_include_water

    def auto_ref(self, tmsp=True):
        if self.nmrdat[self.s][self.e].dim == 1:
            #self.nmrdat[self.s][self.e].ref = 'auto'
            self.nmrdat[self.s][self.e].auto_ref(tmsp)
            # self.nmrdat[self.s][self.e].setRef(np.array([0.0]), np.array([14836]))

        else:
            self.nmrdat[self.s][self.e].auto_ref()
            # ref_shifts                = np.array([1.33, 19.3])
            # ref_points                = np.array([262, 2150])
            # self.nmrdat[self.s][self.e].setRef(ref_shifts, ref_points)

        return "Finished auto_ref"
        # end auto_ref

    def auto_ref_all(self, tmsp=True):
        n_exp = len(self.nmrdat[self.s])
        orig_exp = self.e
        for k in range(n_exp):
            self.e = k
            self.auto_ref(tmsp)

        self.e = orig_exp
        return "Finished auto_ref_all"
        # end auto_ref_all

    def baseline1d(self):
        if self.nmrdat[self.s][self.e].dim == 1:
            if len(self.nmrdat[self.s][self.e].spline_baseline.baseline_points) == 0:
                self.nmrdat[self.s][self.e].baseline1d()

        # end baseline1d

    def baseline1d_all(self):
        orig_exp = self.e
        for k in range(len(self.nmrdat[self.s])):
            self.e = k
            self.baseline1d()

        self.e = orig_exp
        # end baseline1d_all

    def bucket_spectra(self):
        self.pp.bucket_points = np.diff(self.nmrdat[self.s][self.e].ppm2points([0.0, self.pp.bucket_ppm]))
        idx1 = np.arange(len(self.nmrdat[self.s][0].ppm1))
        idx2 = idx1[::int(self.pp.bucket_points[0])]
        idx2 = np.append(idx2, len(idx1))
        ppm = np.array([])
        for k in range(len(idx2) - 1):
            ppm = np.append(ppm, np.mean(self.nmrdat[self.s][0].ppm1[idx2[k]:idx2[k + 1]]))

        for k in range(len(self.nmrdat[self.s])):
            spc = np.array([])
            for l in range(len(idx2) - 1):
                spc = np.append(spc,
                                np.sum(self.nmrdat[self.s][k].spc[0][idx2[l]:idx2[l + 1]].real / self.pp.bucket_points))

            self.nmrdat[self.s][k].ppm1 = ppm
            self.nmrdat[self.s][k].spc = np.resize(self.nmrdat[self.s][k].spc, (1, len(spc)))
            self.nmrdat[self.s][k].spc[0] = np.copy(spc)

        self.pp.bucket_points = 1
        # end bucket_spectra

    def check_file(self, data_set_name=''):
        if len(data_set_name) == 0:
            msg = 'file name empty'
            return msg

        nda = self.load2(data_set_name)
        ds = []
        exps = []
        for k in range(len(nda)):
            exps.append([])
            for l in range(len(nda[k])):
                if len(nda[k][k].spc[0]) == 0 or len(nda[k][l].fid[0]) == 0:
                    ds.append(k)
                    exps[k].append(l)

        ds = list(set(ds))
        msg = ''
        for k in range(len(ds)):
            for l in range(len(exps[ds[k]])):
                msg += f'File: {data_set_name}, Dataset: {ds[k] + 1}, experiment: {exps[ds[k]][l] + 1} not a valid dataset\n'

        return msg
        # end check_file

    def clear(self):
        self.nmrdat = [[]]
        self.s = 0
        self.e = -1
        return "Workspace cleared"
        # end clear

    def compress_buckets(self):
        if len(self.pp.compress_start) != len(self.pp.compress_end):
            return

        for k in range(len(self.pp.compress_start)):
            idx = np.where((self.nmrdat[self.s][0].ppm1 > self.pp.compress_start[k]) & (
                    self.nmrdat[self.s][0].ppm1 < self.pp.compress_end[k]))
            self.deselect[idx] = np.ones(len(idx))
            self.deselect[int(np.round(np.mean(idx)))] = 0
            # print(idx)
            for l in range(len(self.nmrdat[self.s])):
                val = np.sum(self.nmrdat[self.s][l].spc[0][idx])
                self.nmrdat[self.s][l].spc[0][idx] = np.zeros(len(idx))
                self.nmrdat[self.s][l].spc[0][int(np.round(np.mean(idx)))] = val

        # end compress_buckets

    def create_titles(self, xls=[], dataset_label='', pos_label='', rack_label='', replace_title=False, excel_name='', autosampler='SampleJet'):
        if len(xls) == 0 or len(dataset_label) == 0 or len(pos_label) == 0 or len(
                excel_name) == 0:
            return

        if len(rack_label) == 0 and autosampler == 'SampleJet':
            return

        if len(self.nmrdat[self.s]) == 0:
            return

        c_dict = {}
        for k in range(len(xls[pos_label])):
            if str(xls[pos_label][k]) != 'nan':
                if not xls[dataset_label][k] in c_dict.keys():
                    c_dict[xls[dataset_label][k]] = {}

                if autosampler == 'SampleJet':
                    c_dict[xls[dataset_label][k]][str(xls[rack_label][k]).replace(' ','') + " " + str(xls[pos_label][k]).replace(' ','')] = k
                else:
                    c_dict[xls[dataset_label][k]][str(xls[pos_label][k]).replace(' ','')] = k

        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].create_title(xls, dataset_label, pos_label, rack_label, replace_title, c_dict,
                                                excel_name, autosampler)

        # end create_titles

    def data_pre_processing(self):
        self.pp.spc_scale = np.ones(len(self.nmrdat[self.s]))
        if not self.nmrdat[self.s][0].projected_j_res:
            self.ft_all()
            if self.nmrdat[self.s][0].proc.autobaseline:
                self.autobaseline1d_all()
            else:
                self.baseline1d_all()

            self.auto_ref_all()
            self.shift_ref()

        else:
            s = self.s
            e = self.e
            self.s = self.nmrdat[s][e].orig_j_res_set
            self.e = self.nmrdat[s][e].orig_j_res_exp
            self.pjres(s + 1, self.nmrdat[s][e].pjres_mode)
            self.s = s
            self.e = e

        if self.pp.export_method == 0 and self.pp.flag_export_data_set == True:
            self.export_data_set('init')

        self.noise_filtering_init()
        self.deselect = np.zeros(len(self.nmrdat[self.s][0].spc[0]))
        self.deselect2 = np.zeros(len(self.nmrdat[self.s][0].spc[0]))
        if self.pp.flag_exclude_region == True:
            self.exclude_region()

        if self.pp.flag_segmental_alignment == True:
            self.segmental_alignment()

        if self.pp.flag_noise_filtering == True:
            self.noise_filtering()

        idx = np.where(self.deselect == 1)
        idx2 = np.where(self.deselect2 == len(self.nmrdat[self.s]))
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].spc[0][idx] = np.zeros(len(idx))
            self.nmrdat[self.s][k].spc[0][idx2] = np.zeros(len(idx2))

        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].spc[0][idx] = np.zeros(len(idx))
            self.nmrdat[self.s][k].spc[0][idx2] = np.zeros(len(idx2))

        if self.pp.flag_bucket_spectra == True:
            self.bucket_spectra()
            if self.pp.export_method == 0 and self.pp.flag_export_data_set == True:
                self.export_data_set('bucket_spectra')

        if self.pp.flag_compress_buckets == True:
            self.compress_buckets()

        if self.pp.flag_scale_spectra == True:
            self.scale_spectra()
            if self.pp.export_method == 0 and self.pp.flag_export_data_set == True:
                if self.pp.scale_pqn:
                    self.export_data_set('pqn_normalised')
                else:
                    self.export_data_set('tsa_normalised')

        spc = np.zeros(len(self.nmrdat[self.s][0].spc[0]))
        n_spc = len(self.nmrdat[self.s])
        for k in range(n_spc):
            spc += self.nmrdat[self.s][k].spc[0].real

        idx = np.where(spc == 0)
        m_val = 0
        for k in range(len(self.nmrdat[self.s])):
            m_val = min(m_val, np.min(self.nmrdat[self.s][k].spc[0].real))

        if m_val < 0 and self.pp.avoid_negative_values:
            for k in range(len(self.nmrdat[self.s])):
                self.nmrdat[self.s][k].spc[0] -= m_val

        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].spc[0][idx] = np.zeros(len(idx))

        if self.pp.avoid_negative_values and self.pp.flag_export_data_set:
            self.export_data_set('avoid_negative_values')

        if self.pp.flag_variance_stabilisation == True:
            self.variance_stabilisation()
            if self.pp.export_method == 0 and self.pp.flag_export_data_set == True:
                self.export_data_set('pareto_scaled')

        if self.pp.flag_export_data_set == True:
            self.export_data_set('finish')

        # end data_pre_processing

    def exclude_region(self):
        if len(self.pp.exclude_start) != len(self.pp.exclude_end):
            return

        for k in range(len(self.pp.exclude_start)):
            idx = np.where((self.nmrdat[self.s][0].ppm1 > self.pp.exclude_start[k]) & (
                    self.nmrdat[self.s][0].ppm1 < self.pp.exclude_end[k]))
            self.deselect[idx] = np.ones(len(idx))

        # end exclude_region

    def export_bruker_1d(self, path_name='', scale_factor=-1):
        if len(path_name) == 0:
            return

        n_exps = len(self.nmrdat[self.s])
        for k in range(n_exps):
            self.nmrdat[self.s][k].export_bruker_1d(path_name, str(10 * (k + 1)), scale_factor)

        # end export_bruker_1d

    def export_hsqc_data(self, excel_name=''):
        n_cols = 7
        if len(excel_name) == 0:
            return

        n_exp = len(self.nmrdat[self.s])
        wb = Workbook()
        wb.remove(wb.active)
        for k in self.nmrdat[self.s][self.e].hsqc.hsqc_data.keys():
            wb.create_sheet(k)
            letters = []
            for s in itertools.islice(self.iter_all_strings(), n_cols * n_exp):
                letters.append(s)

            for l in range(n_exp):
                wb[k][letters[0 + l * n_cols] + "1"] = "Dataset." + str(l)
                wb[k][letters[1 + l * n_cols] + "1"] = "Experiment." + str(l)
                wb[k][letters[2 + l * n_cols] + "1"] = "Remote." + str(l)
                wb[k][letters[3 + l * n_cols] + "1"] = "Multiplet." + str(l)
                wb[k][letters[4 + l * n_cols] + "1"] = "Percentages." + str(l)
                wb[k][letters[5 + l * n_cols] + "1"] = "Intensity." + str(l)
                wb[k][letters[6 + l * n_cols] + "1"] = "HSQC." + str(l)
                wb[k][letters[0 + l * n_cols] + "2"] = str(self.s + 1)
                wb[k][letters[1 + l * n_cols] + "2"] = str(l + 1)
                wb[k][letters[6 + l * n_cols] + "2"] = str(self.nmrdat[self.s][l].hsqc.hsqc_data[k].hsqc).replace('[',
                                                                                                                  '').replace(
                    ']', '')
                wb[k][letters[6 + l * n_cols] + "3"] = "n_bonds: " + str(
                    self.nmrdat[self.s][l].hsqc.hsqc_data[k].n_bonds)
                offset = 0
                for m in range(len(self.nmrdat[self.s][l].hsqc.hsqc_data[k].spin_systems)):
                    wb[k][letters[5 + l * n_cols] + str(offset + 2)] = str(
                        self.nmrdat[self.s][l].hsqc.hsqc_data[k].intensities[m])
                    for n in range(len(self.nmrdat[self.s][l].hsqc.hsqc_data[k].spin_systems[m]['contribution'])):
                        wb[k][letters[2 + l * n_cols] + str(offset + 2)] = str(
                            self.nmrdat[self.s][l].hsqc.hsqc_data[k].spin_systems[m]['c13_idx'][0]).replace('[',
                                                                                                            '').replace(
                            ']', '')
                        wb[k][letters[3 + l * n_cols] + str(offset + 2)] = str(
                            self.nmrdat[self.s][l].hsqc.hsqc_data[k].spin_systems[m]['c13_idx'][n]).replace('[',
                                                                                                            '').replace(
                            ']', '')
                        wb[k][letters[4 + l * n_cols] + str(offset + 2)] = str(
                            self.nmrdat[self.s][l].hsqc.hsqc_data[k].spin_systems[m]['contribution'][n])
                        offset += 1

        wb.save(excel_name)

    def export_data_set(self, cmd_name='finish'):
        if self.pp.export_method == 0:
            if os.path.isdir(self.pp.export_excel_path) is False:
                os.makedirs(self.pp.export_excel_path)

            f_name = os.path.join(self.pp.export_excel_path, self.pp.export_excel)
            if cmd_name == 'init':
                self.wb = Workbook()
                self.wb.remove(self.wb.active)
                ws_bucket_spectra = self.wb.create_sheet("bucket_spectra")
                if self.pp.scale_pqn:
                    ws_pqn_normalised = self.wb.create_sheet("pqn_normalised")
                else:
                    ws_pqn_normalised = self.wb.create_sheet("tsa_normalised")

                ws_avoid_negative_values = self.wb.create_sheet("avoid_negative_values")
                ws_pareto_scaled = self.wb.create_sheet("pareto_scaled")
                ws_sample_meta = self.wb.create_sheet("sample_meta")
                ws_variable_meta = self.wb.create_sheet("variable_meta")
                self.wb["variable_meta"]["A1"] = 'id'
                self.wb["variable_meta"]["B1"] = 'ppm'
                self.wb["variable_meta"]["C1"] = 'bin'
                self.wb["variable_meta"]["D1"] = 'ppm_min'
                self.wb["variable_meta"]["E1"] = 'ppm_max'
            else:
                if cmd_name != 'finish':
                    spc = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
                    npts = len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0])
                    n_spc = len(self.pp.plot_select)
                    for k in range(n_spc):
                        spc += self.nmrdat[self.s][self.pp.plot_select[k]].spc[0].real

                    deselect = np.zeros(npts)
                    idx = np.where(spc == 0)
                    deselect[idx] = np.ones(len(idx))
                    select = np.where(deselect == 0)
                    if self.pp.export_samples_in_rows_cols == 0:  # samples in rows
                        col_string = []
                        for s in itertools.islice(self.iter_all_strings(), len(select[0]) + 2):
                            col_string.append(s)

                        self.wb[cmd_name]['A1'] = 'Name'
                        self.wb[cmd_name]['B1'] = 'Class / ppm -->'
                        for k in range(len(select[0])):
                            self.wb[cmd_name][col_string[k + 2] + '1'] = str(
                                self.nmrdat[self.s][self.pp.plot_select[0]].ppm1[select[0][k]])

                        for k in range(len(self.pp.plot_select)):
                            dse = os.path.split(self.nmrdat[self.s][self.pp.plot_select[k]].orig_data_set)
                            ds = os.path.split(dse[0])
                            self.wb[cmd_name]['A' + str(k + 2)] = ds[1] + " " + dse[1]
                            self.wb[cmd_name]['B' + str(k + 2)] = str(self.pp.class_select[self.pp.plot_select[k]])
                            for l in range(len(select[0])):
                                self.wb[cmd_name][col_string[l + 2] + str(k + 2)] = str(
                                    self.nmrdat[self.s][self.pp.plot_select[k]].spc[0][select[0][l]].real)



                    else:  # samples in cols
                        self.wb[cmd_name]['A1'] = ''
                        col_string = []
                        for s in itertools.islice(self.iter_all_strings(), n_spc + 1):
                            col_string.append(s)

                        for k in range(len(self.pp.plot_select)):
                            dse = os.path.split(self.nmrdat[self.s][self.pp.plot_select[k]].orig_data_set)
                            ds = os.path.split(dse[0])
                            self.wb[cmd_name][col_string[k + 1] + '1'] = ds[1] + " " + dse[1]

                        ppm_vect = self.nmrdat[self.s][self.pp.plot_select[0]].ppm1
                        bin_range = np.linspace(1, len(ppm_vect), len(ppm_vect), dtype=int)
                        spc_selected = np.where(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0] != 0)[0]
                        for k in range(len(spc_selected)):
                            p_str1 = "B{0:0=3d}".format(bin_range[spc_selected[k]])
                            p_str2 = "P{:.3f}".format(ppm_vect[spc_selected[k]])
                            self.wb[cmd_name]["A" + str(k + 2)] = p_str1 + p_str2
                            for l in range(len(self.pp.plot_select)):
                                self.wb[cmd_name][col_string[l + 1] + str(k + 2)] = \
                                self.nmrdat[self.s][self.pp.plot_select[l]].spc[0][spc_selected[k]].real

                else:
                    categories = {}
                    title = self.nmrdat[self.s][self.e].title
                    idx = 0
                    while idx > -1:
                        idx = title.find('\n')
                        if idx > -1:
                            idx2 = title.find(':')
                            categories[title[:idx2].strip()] = []
                            title = title[idx + 1:]

                    if self.pp.scale_pqn:
                        categories['pqn_coeff'] = []
                    else:
                        categories['tsa_coeff'] = []

                    excel_cols = []
                    for ss in itertools.islice(self.iter_all_strings(), len(categories)):
                        excel_cols.append(ss)

                    ctr = 0
                    for ss in categories.keys():
                        self.wb["sample_meta"][excel_cols[ctr] + '1'] = ss
                        ctr += 1

                    for k in range(len(self.pp.plot_select)):
                        for ss in categories.keys():
                            if ss == "pqn_coeff" or ss == "tsa_coeff":
                                categories[ss].append(self.pp.spc_scale[self.pp.plot_select[k]])
                            else:
                                idx = self.nmrdat[self.s][self.pp.plot_select[k]].title.find(ss)
                                tmp_string = self.nmrdat[self.s][self.pp.plot_select[k]].title[idx:]
                                idx2 = tmp_string.find(':')
                                idx3 = tmp_string.find('\n')
                                categories[ss].append(tmp_string[idx2 + 1:idx3])

                    ctr2 = 2
                    for k in range(len(self.pp.plot_select)):
                        ctr = 0
                        for ss in categories.keys():
                            self.wb["sample_meta"][excel_cols[ctr] + str(ctr2)] = categories[ss][k]
                            ctr += 1

                        ctr2 += 1

                    ppm_vect = self.nmrdat[self.s][self.pp.plot_select[0]].ppm1
                    delta_ppm = 0.5 * (ppm_vect[0] - ppm_vect[1])
                    bin_range = np.linspace(1, len(ppm_vect), len(ppm_vect), dtype=int)
                    spc_selected = np.where(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0] != 0)[0]
                    for k in range(len(spc_selected)):
                        p_str1 = "B{0:0=3d}".format(bin_range[spc_selected[k]])
                        p_str2 = "P{:.3f}".format(ppm_vect[spc_selected[k]])
                        self.wb["variable_meta"]["A" + str(k + 2)] = p_str1 + p_str2
                        self.wb["variable_meta"]["B" + str(k + 2)] = ppm_vect[spc_selected[k]]
                        self.wb["variable_meta"]["C" + str(k + 2)] = bin_range[spc_selected[k]]
                        if k > 0 and k < len(spc_selected) - 1:
                            if spc_selected[k] - spc_selected[k - 1] > 1 and spc_selected[k + 1] - spc_selected[k] > 1:
                                self.wb["variable_meta"]["D" + str(k + 2)] = ppm_vect[
                                                                                 spc_selected[k + 1] - 1] - delta_ppm
                                self.wb["variable_meta"]["E" + str(k + 2)] = ppm_vect[
                                                                                 spc_selected[k - 1] + 1] + delta_ppm
                            else:
                                self.wb["variable_meta"]["D" + str(k + 2)] = ppm_vect[spc_selected[k]] - delta_ppm
                                self.wb["variable_meta"]["E" + str(k + 2)] = ppm_vect[spc_selected[k]] + delta_ppm
                        else:
                            self.wb["variable_meta"]["D" + str(k + 2)] = ppm_vect[spc_selected[k]] - delta_ppm
                            self.wb["variable_meta"]["E" + str(k + 2)] = ppm_vect[spc_selected[k]] + delta_ppm

                    self.wb.save(f_name)
                    self.wb = []

        elif self.pp.export_method == 1:
            print("export CSV format")
            if os.path.isdir(self.pp.export_path_name) is False:
                os.makedirs(self.pp.export_path_name)

            f_name = os.path.join(self.pp.export_path_name, self.pp.export_file_name)
            f = open(f_name, 'w')
            if self.pp.export_delimiter_tab:
                delim = '\t'
            else:
                delim = self.pp.export_character

            if len(self.pp.plot_select) == 0:
                self.pp.plot_select = [0]

            spc = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            for k in range(len(self.pp.plot_select)):
                spc += self.nmrdat[self.s][self.pp.plot_select[k]].spc[0].real

            deselect = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            idx = np.where(spc == 0)
            deselect[idx] = np.ones(len(idx))
            if self.pp.export_samples_in_rows_cols == 0:  # samples in rows
                f.write("ppm" + delim + " ")
                for k in range(len(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1)):
                    if deselect[k] == 0:
                        f.write(delim + str(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1[k]))

                f.write("\n")
                for k in range(len(self.pp.plot_select)):
                    dse = os.path.split(self.nmrdat[self.s][self.pp.plot_select[k]].orig_data_set)
                    ds = os.path.split(dse[0])
                    f.write(ds[1] + " " + dse[1] + delim + self.pp.class_select[self.pp.plot_select[k]])
                    for l in range(len(self.nmrdat[self.s][self.pp.plot_select[k]].spc[0])):
                        if deselect[l] == 0:
                            f.write(delim + str(self.nmrdat[self.s][self.pp.plot_select[k]].spc[0][l].real))

                    f.write("\n")

            else:  # samples in cols
                f.write("ppm")
                for k in range(len(self.pp.plot_select)):
                    dse = os.path.split(self.nmrdat[self.s][self.pp.plot_select[k]].orig_data_set)
                    ds = os.path.split(dse[0])
                    f.write(delim + ds[1] + " " + dse[1])

                f.write("\n")
                f.write(" ")
                for k in range(len(self.pp.plot_select)):
                    f.write(delim + self.pp.class_select[self.pp.plot_select[k]])

                f.write("\n")
                for k in range(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0])):
                    if deselect[k] == 0:
                        f.write(str(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1[k]))
                        for l in range(len(self.pp.plot_select)):
                            f.write(delim + str(self.nmrdat[self.s][self.pp.plot_select[l]].spc[0][k].real))

                        f.write("\n")

            f.close()

        elif self.pp.export_method == 2:
            print("export MetaboAnalyst")
            if not os.path.isdir(self.pp.export_metabo_analyst_path):
                os.makedirs(self.pp.export_metabo_analyst_path)

            f_name = os.path.join(self.pp.export_metabo_analyst_path, self.pp.export_metabo_analyst)
            f = open(f_name, 'w')
            delim = ','
            spc = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            for k in range(len(self.pp.plot_select)):
                spc += self.nmrdat[self.s][self.pp.plot_select[k]].spc[0].real

            deselect = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            idx = np.where(spc == 0)
            deselect[idx] = np.ones(len(idx))
            f.write("Sample" + delim + " Class")
            for k in range(len(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1)):
                if deselect[k] == 0:
                    f.write(delim + " Bin." + str(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1[k]))

            f.write("\n")
            for k in range(len(self.pp.plot_select)):
                dse = os.path.split(self.nmrdat[self.s][self.pp.plot_select[k]].orig_data_set)
                ds = os.path.split(dse[0])
                f.write(ds[1] + " " + dse[1] + delim + " " + self.pp.class_select[k])
                for l in range(len(self.nmrdat[self.s][k].spc[self.pp.plot_select[0]])):
                    if deselect[l] == 0:
                        f.write(delim + " " + str(self.nmrdat[self.s][k].spc[self.pp.plot_select[0]][l].real))

                f.write("\n")

            f.close()

        elif self.pp.export_method == 3:
            print("export rDolphin")
            if os.path.isdir(self.pp.export_r_dolphin_path) is False:
                os.makedirs(self.pp.export_r_dolphin_path)

            f_name = os.path.join(self.pp.export_r_dolphin_path, self.pp.export_r_dolphin)
            f = open(f_name, 'w')
            delim = ','
            spc = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            for k in range(len(self.pp.plot_select)):
                spc += self.nmrdat[self.s][self.pp.plot_select[k]].spc[0].real

            deselect = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            idx = np.where(spc == 0)
            deselect[idx] = np.ones(len(idx))
            f.write(str(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1[0]))
            for k in range(1, len(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1)):
                if deselect[k] == 0:
                    f.write(delim + str(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1[k]))

            f.write("\n")
            for k in range(len(self.pp.plot_select)):
                dse = os.path.split(self.nmrdat[self.s][self.pp.plot_select[k]].orig_data_set)
                ds = os.path.split(dse[0])
                f.write(str(self.nmrdat[self.s][self.pp.plot_select[k]].spc[0][0].real))
                for l in range(1, len(self.nmrdat[self.s][self.pp.plot_select[k]].spc[0])):
                    if deselect[l] == 0:
                        f.write(delim + str(self.nmrdat[self.s][self.pp.plot_select[k]].spc[0][l].real))

                f.write("\n")

            f.close()
            f_name = os.path.join(self.pp.export_r_dolphin_path, "Parameters.csv")
            f = open(f_name, 'w')
            f.write("Parameter,Value\n")
            f.write("nmr folder path,\n")
            f.write("1D data index,\n")
            f.write("proc_no,\n")
            f.write("spectra dataset path (csv format)," + self.pp.export_r_dolphin_path + "/" +
                    self.pp.export_r_dolphin + "\n")
            f.write("Metadata path (csv format)," + self.pp.export_r_dolphin_path + "/Metadata.csv\n")
            f.write("ROI patters file," + self.pp.export_r_dolphin_path + "/ROI_profile.csv\n")
            f.write("Normalization (0=No;1=Eretic; 2=TSP; 3=Creatinine; 4=Spectra Sum; 5=PQN),2\n")
            f.write("Alignment (0=No;1=Glucose; 2=TSP; 3=Formate),2\n")
            f.write("Suppression,12-9.5;6.1-5.6;5.1-4.5\n")
            f.write("Spectrometer Frequency (MHz)," + str(self.nmrdat[self.s][self.pp.plot_select[0]].acq.sfo1) + "\n")
            f.write("Bucket resolution," + str(self.pp.bucket_ppm) + "\n")
            f.write("Biofluid,Urine\n")
            f.write("2D-Path,\n")
            f.write("Specific program parameters,\n")
            f.close()
            f_name = os.path.join(self.pp.export_r_dolphin_path, "Metadata.csv")
            f = open(f_name, 'w')
            f.write("Sample,Individual,Sample Type\n")
            for k in range(len(self.pp.plot_select)):
                f.write(os.path.split(self.nmrdat[self.s][self.pp.plot_select[k]].orig_data_set)[1] + "," + str(
                    k + 1) + ",1\n")

            f.close()

        elif self.pp.export_method == 4:
            print("export Batman")
            if os.path.isdir(self.pp.export_batman_path) is False:
                os.makedirs(self.pp.export_batman_path)

            f_name = os.path.join(self.pp.export_batman_path, self.pp.export_batman)
            f = open(f_name, 'w')
            delim = '\t'
            spc = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            for k in range(len(self.pp.plot_select)):
                spc += self.nmrdat[self.s][self.pp.plot_select[k]].spc[0].real

            deselect = np.zeros(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0]))
            idx = np.where(spc == 0)
            deselect[idx] = np.ones(len(idx))
            f.write("ppm")
            for k in range(len(self.pp.plot_select)):
                f.write(delim + "X" + str(k + 1))

            f.write("\n")
            for k in range(len(self.nmrdat[self.s][self.pp.plot_select[0]].spc[0])):
                if deselect[k] == 0:
                    f.write(str(self.nmrdat[self.s][self.pp.plot_select[0]].ppm1[k]))
                    for l in range(len(self.pp.plot_select)):
                        f.write(delim + str(self.nmrdat[self.s][self.pp.plot_select[l]].spc[0][k].real))

                    f.write("\n")

            f.close()

        elif self.pp.export_method == 5:
            print("export Bruker")
            if os.path.isdir(self.pp.export_bruker_path + os.sep + self.pp.export_bruker) is False:
                os.makedirs(self.pp.export_bruker_path + os.sep + self.pp.export_bruker)
            else:
                shutil.rmtree(self.pp.export_bruker_path + os.sep + self.pp.export_bruker)

            m_max = 0
            for k in range(len(self.pp.plot_select)):
                m_max = max(np.max(self.nmrdat[self.s][self.pp.plot_select[k]].spc[0].real), m_max)

            scale_factor = 2 * m_max / 2147483647
            for k in range(len(self.pp.plot_select)):
                self.nmrdat[self.s][self.pp.plot_select[k]].export_bruker_1d(
                    self.pp.export_bruker_path + os.sep + self.pp.export_bruker,
                    str(k + 1), scale_factor)

    # end export_data_set

    def fitlw(self, ppm=0.0, message=True):
        self.nmrdat[self.s][self.e].fitlw(ppm=ppm)
        if message:
            print(f'Fitting linewidth@{ppm:4.2f}ppm for dataset: {self.s + 1}, experiment: {self.e + 1:03d}, peak linewidth: {self.nmrdat[self.s][self.e].tmsp_linewidth:4.3f} Hz, offset: {self.nmrdat[self.s][self.e].tmsp_offset:+4.4f}, M0: {self.nmrdat[self.s][self.e].tmsp_m0:4.2f}')

    # end fit_tmsp

    def fitlw_all(self, ppm=0.0, message=True):
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].fitlw(ppm=ppm)
            if message:
                print(f'Fitting linewidth@{ppm:4.2f}ppm for dataset: {self.s + 1}, experiment: {k + 1:03d}, peak linewidth: {self.nmrdat[self.s][k].tmsp_linewidth:4.3f} Hz, offset: {self.nmrdat[self.s][k].tmsp_offset:+4.4f}, M0: {self.nmrdat[self.s][k].tmsp_m0:4.2f}')

        # end fit_tmsp_all

    def peakw(self, ppm=0.0, message=True):
        self.nmrdat[self.s][self.e].peakw(ppm=ppm)
        if message:
            print(f'Peak width@{ppm:4.2f}ppm for dataset: {self.s + 1}, experiment: {self.e + 1:03d}, peak linewidth: {self.nmrdat[self.s][self.e].tmsp_linewidth:4.3f} Hz')
        # end peakw_tmsp

    def peakw_all(self, ppm=0.0, message=True):
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].peakw(ppm=ppm)
            if message:
                print(f'Peak width@{ppm:4.2f}ppm for dataset: {self.s + 1}, experiment: {k + 1:03d}, peak linewidth: {self.nmrdat[self.s][k].tmsp_linewidth:4.3f} Hz')
        # end peakw_tmsp_all

    def ft(self):
        if self.nmrdat[self.s][self.e].dim == 1:
            self.nmrdat[self.s][self.e].proc_spc1d()

        else:
            self.nmrdat[self.s][self.e].proc_spc()

        # end ft

    def ft_all(self):
        n_exp = len(self.nmrdat[self.s])
        orig_exp = self.e
        for k in range(n_exp):
            self.e = k
            self.ft()
        #
        self.e = orig_exp
        return "Finished ft_all"
        # end ft_all

    def help(self):  # pragme: no cover
        url = "https://ludwigc.github.io/metabolabpy"
        webbrowser.open(url, new=2)
        # end help

    def init_print_colours(self):
        self.cf.read_config()
        if self.cf.print_light_mode == False:
            self.int1 = 1.0
            self.int2 = 0.6
            self.int3 = 0.3
            self.int4 = 0.8
        else:
            self.int1 = 0.4
            self.int2 = 0.8
            self.int3 = 0.5
            self.int4 = 0.1

        int1 = self.int1
        int2 = self.int2
        int3 = self.int3
        int4 = self.int4
        light_neg_diff = 0.5
        dark_neg_diff = 0.2
        if self.cf.print_light_mode == False:
            self.print_colours = [(int4, int4, int1),
                                  (0.0, int1, int1),
                                  (int1, 0.0, int1),
                                  (int2, int2, int1),
                                  (int1, int2, int2),
                                  (int2, int1, int2),
                                  (int1, int1, int3),
                                  (int2, int3, int3),
                                  (int3, int2, int3),
                                  (int3, int2, int2),
                                  (int2, int2, int3),
                                  (int2, int3, int2)]

            int1 = self.int1 - dark_neg_diff
            int2 = self.int2 - dark_neg_diff
            int3 = self.int3 - dark_neg_diff
            int1 = max(int1, 0.0)
            int2 = max(int2, 0.0)
            int3 = max(int3, 0.0)
            self.print_neg_colours = [(int2, int2, int1),
                                      (0.0, int1, int1),
                                      (int1, 0.0, int1),
                                      (int2, int2, int1),
                                      (int1, int2, int2),
                                      (int2, int1, int2),
                                      (int1, int1, int3),
                                      (int2, int3, int3),
                                      (int3, int2, int3),
                                      (int3, int2, int2),
                                      (int2, int2, int3),
                                      (int2, int3, int2)]

        else:
            self.print_colours = [(0.0, 0.0, int1),
                                  (int1, 0.0, 0.0),
                                  (0.0, int1, 0.0),
                                  (0.0, int1, int1),
                                  (int1, int1, 0.0),
                                  (int1, 0.0, int1),
                                  (int3, int3, int2),
                                  (int2, int3, int3),
                                  (int3, int2, int3),
                                  (int3, int2, int2),
                                  (int2, int2, int3),
                                  (int2, int3, int2)]

            int1 = self.int1 + light_neg_diff
            int2 = self.int2 + light_neg_diff
            int3 = self.int3 + light_neg_diff
            int1 = min(int1, 1.0)
            int2 = min(int2, 1.0)
            int3 = min(int3, 1.0)
            self.print_neg_colours = [(light_neg_diff, light_neg_diff, int1),
                                      (int1, light_neg_diff, light_neg_diff),
                                      (light_neg_diff, int1, light_neg_diff),
                                      (light_neg_diff, int1, int1),
                                      (int1, int1, light_neg_diff),
                                      (int1, light_neg_diff, int1),
                                      (int3, int3, int2),
                                      (int2, int3, int3),
                                      (int3, int2, int3),
                                      (int3, int2, int2),
                                      (int2, int2, int3),
                                      (int2, int3, int2)]

        self.std_pos_col1 = (self.cf.pos_col10, self.cf.pos_col11, self.cf.pos_col12)
        self.std_neg_col1 = (self.cf.neg_col10, self.cf.neg_col11, self.cf.neg_col12)
        self.std_pos_col2 = (self.cf.pos_col20, self.cf.pos_col21, self.cf.pos_col22)
        self.std_neg_col2 = (self.cf.neg_col20, self.cf.neg_col21, self.cf.neg_col22)
        if self.cf.print_light_mode == False:
            self.print_pos_col_rgb = self.std_pos_col2
            self.print_neg_col_rgb = self.std_neg_col2
            self.print_background_colour = (42 / 255, 42 / 255, 42 / 255)
            self.print_foreground_colour = (255 / 255, 255 / 255, 255 / 255)
        else:
            self.print_pos_col_rgb = self.std_pos_col1
            self.print_neg_col_rgb = self.std_neg_col1
            self.print_background_colour = (255 / 255, 255 / 255, 255 / 255)
            self.print_foreground_colour = (0 / 255, 0 / 255, 0 / 255)

        # end init_print_colours

    def iter_all_strings(self):
        for size in itertools.count(1):
            for s in itertools.product(ascii_uppercase, repeat=size):
                yield "".join(s)

        # end iter_all_strings

    def load(self, data_set_name):
        data_sets = np.array([])
        l_dir = os.listdir(data_set_name)
        cur_data_not_found = True
        for k in range(len(l_dir)):
            if l_dir[k] == 'curPars.dat':
                cur_data_not_found = False

        if cur_data_not_found:
            print('no curPars.dat!')
            return

        f_name = os.path.join(data_set_name, 'curPars.dat')
        f = open(f_name, 'rb')
        cur_pars = pickle.load(f)
        f.close()
        self.file_format_version = cur_pars[0]
        self.s = cur_pars[1]
        self.e = cur_pars[2]
        # self.pp = cur_pars[3]
        c = cur_pars[3]
        for k in self.pp.__dict__.keys():
            if k != 'cf':
                k2 = k
                str_idx = k2.find('_')
                while str_idx != -1:
                    str_letter = k2[str_idx + 1]
                    k2 = k2.replace(k2[str_idx:str_idx + 2], str_letter.upper())
                    str_idx = k2.find('_')

                if hasattr(c, k):
                    exec('self.pp.' + k + '=c.' + k)

                if hasattr(c, k2):
                    exec('self.pp.' + k + '=c.' + k2)

        self.deselect = cur_pars[4]
        self.deselect2 = cur_pars[5]
        self.cmd_buffer = cur_pars[6]
        self.cmd_idx = cur_pars[7]
        self.script = cur_pars[8]
        self.console = cur_pars[9]
        try:
            self.tmsp_conc = cur_pars[10]
        except:
            self.tmsp_conc = 0.5

        for k in range(len(l_dir)):
            if os.path.isdir(os.path.join(data_set_name, l_dir[k])):
                data_sets = np.append(data_sets, l_dir[k])

        data_sets = np.sort(data_sets)
        data_set_exps = []
        for k in range(len(data_sets)):
            dir_name = os.listdir(os.path.join(data_set_name, data_sets[k]))
            data_exps = np.array([])
            for l in range(len(dir_name)):
                if os.path.isdir(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l])):
                    if (os.path.isfile(
                            os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                         'titleFile.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'acqusText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'acqu2sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'acqu3sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'procsText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'proc2sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'proc3sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'nmrDataSet.dat'))):
                        data_exps = np.append(data_exps, dir_name[l])

            data_exps2 = []
            for k in range(len(data_exps)):
                try:
                    data_exps2.append(int(data_exps[k]))
                except:
                    pass

            data_exps2.sort()
            data_exps = list(map(str, data_exps2))
            data_set_exps.append(data_exps)

        self.nmrdat = []
        for k in range(len(data_set_exps)):
            self.nmrdat.append([])
            for l in range(len(data_set_exps[k])):
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'nmrDataSet.dat')
                f = open(f_name, 'rb')
                n = pickle.load(f)
                f.close()
                if hasattr(n, 'ver'):
                    vver = ''
                else:
                    vver = '0.1'

                nd2 = nd.NmrData()
                # nd2.file_version = n.nmrdat[0][0].ver
                for kk in nd2.__dict__.keys():
                    if kk != 'acq' and kk != 'proc' and kk != 'display' and kk != 'apc' and kk != 'hsqc':
                        kk2 = kk
                        str_idx = kk2.find('_')
                        while str_idx != -1:
                            str_letter = kk2[str_idx + 1]
                            kk2 = kk2.replace(kk2[str_idx:str_idx + 2], str_letter.upper())
                            str_idx = kk2.find('_')

                        if hasattr(n, kk):
                            exec('nd2.' + kk + '=n.' + kk)

                        if hasattr(n, kk2):
                            exec('nd2.' + kk + '=n.' + kk2)

                    elif kk == 'acq':
                        if hasattr(n, kk):
                            a = n.acq
                            aq = nd2.acq
                            for kkk in aq.__dict__.keys():
                                if kkk != 'reg_ex':
                                    kkk2 = kkk
                                    str_idx = kkk2.find('_')
                                    while str_idx != -1:
                                        str_letter = kkk2[str_idx + 1]
                                        kkk2 = kkk2.replace(kkk2[str_idx:str_idx + 2], str_letter.upper())
                                        str_idx = kkk2.find('_')

                                    if hasattr(a, kkk):
                                        exec('aq.' + kkk + '=a.' + kkk)

                                    if hasattr(a, kkk2):
                                        exec('aq.' + kkk + '=a.' + kkk2)

                                # else:
                                #    try:
                                #        r = a.reg_ex
                                #    except:
                                #        r = a.regEx
                                #
                                #    re = aq.reg_ex
                                #    for kkkk in re.__dict__.keys():
                                #        if hasattr(r, kkkk):
                                #            exec('re.' + kkkk + '=r.' + kkkk)
                                #
                                #    aq.reg_ex = re

                            nd2.acq = aq

                    elif kk == 'proc':
                        if hasattr(n, kk):
                            p = n.proc
                            pc = nd2.proc
                            for kkk in pc.__dict__.keys():
                                if kkk != 'reg_ex':
                                    kkk2 = kkk
                                    str_idx = kkk2.find('_')
                                    while str_idx != -1:
                                        str_letter = kkk2[str_idx + 1]
                                        kkk2 = kkk2.replace(kkk2[str_idx:str_idx + 2], str_letter.upper())
                                        str_idx = kkk2.find('_')

                                    if hasattr(p, kkk):
                                        exec('pc.' + kkk + '=p.' + kkk)

                                    if hasattr(p, kkk2):
                                        exec('pc.' + kkk + '=p.' + kkk2)

                                else:
                                    # try:
                                    #    r = p.reg_ex
                                    # except:
                                    #    r = p.regEx
                                    #
                                    re = pc.reg_ex
                                    # for kkkk in re.__dict__.keys():
                                    #    if hasattr(r, kkkk):
                                    #        exec('re.' + kkkk + '=r.' + kkkk)
                                    #
                                    # pc.reg_ex = re

                            nd2.proc = pc

                    elif kk == 'display':
                        if hasattr(n, kk):
                            d = n.display
                            dp = nd2.display
                            for kkk in dp.__dict__.keys():
                                if hasattr(d, kkk):
                                    exec('dp.' + kkk + '=d.' + kkk)

                            nd2.display = dp

                    elif kk == 'apc':
                        if hasattr(n, kk):
                            ab = n.apc
                            ac = nd2.apc
                            for kkk in ac.__dict__.keys():
                                if hasattr(ab, kkk):
                                    exec('ac.' + kkk + '=ab.' + kkk)

                            nd2.apc = ac

                    elif kk == 'hsqc':
                        if hasattr(n, kk):
                            h = n.hsqc
                            hp = nd2.hsqc
                            for kkk in hp.__dict__.keys():
                                if hasattr(h, kkk):
                                    # if kkk != 'hsqc_data':
                                    exec('hp.' + kkk + '=h.' + kkk)
                                    #
                                    # else:
                                    #    hd = h.hsqc_data
                                    #    hdp = hp.hsqc_data
                                    #    for kkkk in hdp.__dict__.keys():
                                    #        if hasattr(hd, kkkk):
                                    #            exec('hdp.' + kkkk + '=hd.' + kkkk)
                                    #
                                    #    hp.hsqc_data = hdp

                            nd2.hsqc = hp

                if vver == '0.1':
                    nd2.ver = '0.1'

                self.nmrdat[k].append(nd2)
                nd2 = []
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'titleFile.txt')
                f = open(f_name, 'r')
                self.nmrdat[k][l].title = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'acqusText.txt')
                f = open(f_name, 'r')
                self.nmrdat[k][l].acq.acqus_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'acqu2sText.txt')
                f = open(f_name, 'r')
                self.nmrdat[k][l].acq.acqu2s_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'acqu3sText.txt')
                f = open(f_name, 'r')
                self.nmrdat[k][l].acq.acqu3s_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'procsText.txt')
                f = open(f_name, 'r')
                self.nmrdat[k][l].proc.procs_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'proc2sText.txt')
                f = open(f_name, 'r')
                self.nmrdat[k][l].proc.proc2s_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'proc3sText.txt')
                f = open(f_name, 'r')
                self.nmrdat[k][l].proc.proc3s_text = f.read()
                f.close()

        self.cf.current_directory = data_set_name
        self.cf.save_config()
        if len(self.nmrdat) - 1 < self.s:
            self.s = 0

        if len(self.nmrdat[self.s]) - 1< self.e:
            self.e = 0

        # end load

    def load2(self, data_set_name):
        data_sets = np.array([])
        l_dir = os.listdir(data_set_name)
        cur_data_not_found = True
        for k in range(len(l_dir)):
            if l_dir[k] == 'curPars.dat':
                cur_data_not_found = False

        if cur_data_not_found:
            return

        f_name = os.path.join(data_set_name, 'curPars.dat')
        f = open(f_name, 'rb')
        cur_pars = pickle.load(f)
        f.close()
        self.file_format_version = cur_pars[0]
        self.s = cur_pars[1]
        self.e = cur_pars[2]
        # self.pp = cur_pars[3]
        c = cur_pars[3]
        for k in self.pp.__dict__.keys():
            if k != 'cf':
                k2 = k
                str_idx = k2.find('_')
                while str_idx != -1:
                    str_letter = k2[str_idx + 1]
                    k2 = k2.replace(k2[str_idx:str_idx + 2], str_letter.upper())
                    str_idx = k2.find('_')

                if hasattr(c, k):
                    exec('self.pp.' + k + '=c.' + k)

                if hasattr(c, k2):
                    exec('self.pp.' + k + '=c.' + k2)

        self.deselect = cur_pars[4]
        self.deselect2 = cur_pars[5]
        self.cmd_buffer = cur_pars[6]
        self.cmd_idx = cur_pars[7]
        self.script = cur_pars[8]
        self.console = cur_pars[9]
        try:
            self.tmsp_conc = cur_pars[10]
        except:
            self.tmsp_conc = 0.5

        for k in range(len(l_dir)):
            if os.path.isdir(os.path.join(data_set_name, l_dir[k])):
                data_sets = np.append(data_sets, l_dir[k])

        data_sets = np.sort(data_sets)
        data_set_exps = []
        for k in range(len(data_sets)):
            dir_name = os.listdir(os.path.join(data_set_name, data_sets[k]))
            data_exps = np.array([])
            for l in range(len(dir_name)):
                if os.path.isdir(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l])):
                    if (os.path.isfile(
                            os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                         'titleFile.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'acqusText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'acqu2sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'acqu3sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'procsText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'proc2sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'proc3sText.txt')) and
                            os.path.isfile(
                                os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), dir_name[l]),
                                             'nmrDataSet.dat'))):
                        data_exps = np.append(data_exps, dir_name[l])

            data_exps2 = []
            for k in range(len(data_exps)):
                try:
                    data_exps2.append(int(data_exps[k]))
                except:
                    pass

            data_exps2.sort()
            data_exps = list(map(str, data_exps2))
            data_set_exps.append(data_exps)

        nda = []
        for k in range(len(data_set_exps)):
            nda.append([])
            for l in range(len(data_set_exps[k])):
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'nmrDataSet.dat')
                f = open(f_name, 'rb')
                n = pickle.load(f)
                f.close()
                if hasattr(n, 'ver'):
                    vver = ''
                else:
                    vver = '0.1'

                nd2 = nd.NmrData()
                # nd2.file_version = n.nmrdat[0][0].ver
                for kk in nd2.__dict__.keys():
                    if kk != 'acq' and kk != 'proc' and kk != 'display' and kk != 'apc' and kk != 'hsqc':
                        kk2 = kk
                        str_idx = kk2.find('_')
                        while str_idx != -1:
                            str_letter = kk2[str_idx + 1]
                            kk2 = kk2.replace(kk2[str_idx:str_idx + 2], str_letter.upper())
                            str_idx = kk2.find('_')

                        if hasattr(n, kk):
                            exec('nd2.' + kk + '=n.' + kk)

                        if hasattr(n, kk2):
                            exec('nd2.' + kk + '=n.' + kk2)

                    elif kk == 'acq':
                        if hasattr(n, kk):
                            a = n.acq
                            aq = nd2.acq
                            for kkk in aq.__dict__.keys():
                                if kkk != 'reg_ex':
                                    kkk2 = kkk
                                    str_idx = kkk2.find('_')
                                    while str_idx != -1:
                                        str_letter = kkk2[str_idx + 1]
                                        kkk2 = kkk2.replace(kkk2[str_idx:str_idx + 2], str_letter.upper())
                                        str_idx = kkk2.find('_')

                                    if hasattr(a, kkk):
                                        exec('aq.' + kkk + '=a.' + kkk)

                                    if hasattr(a, kkk2):
                                        exec('aq.' + kkk + '=a.' + kkk2)

                                # else:
                                #    try:
                                #        r = a.reg_ex
                                #    except:
                                #        r = a.regEx
                                #
                                #    re = aq.reg_ex
                                #    for kkkk in re.__dict__.keys():
                                #        if hasattr(r, kkkk):
                                #            exec('re.' + kkkk + '=r.' + kkkk)
                                #
                                #    aq.reg_ex = re

                            nd2.acq = aq

                    elif kk == 'proc':
                        if hasattr(n, kk):
                            p = n.proc
                            pc = nd2.proc
                            for kkk in pc.__dict__.keys():
                                if kkk != 'reg_ex':
                                    kkk2 = kkk
                                    str_idx = kkk2.find('_')
                                    while str_idx != -1:
                                        str_letter = kkk2[str_idx + 1]
                                        kkk2 = kkk2.replace(kkk2[str_idx:str_idx + 2], str_letter.upper())
                                        str_idx = kkk2.find('_')

                                    if hasattr(p, kkk):
                                        exec('pc.' + kkk + '=p.' + kkk)

                                    if hasattr(p, kkk2):
                                        exec('pc.' + kkk + '=p.' + kkk2)

                                else:
                                    # try:
                                    #    r = p.reg_ex
                                    # except:
                                    #    r = p.regEx
                                    #
                                    re = pc.reg_ex
                                    # for kkkk in re.__dict__.keys():
                                    #    if hasattr(r, kkkk):
                                    #        exec('re.' + kkkk + '=r.' + kkkk)
                                    #
                                    # pc.reg_ex = re

                            nd2.proc = pc

                    elif kk == 'display':
                        if hasattr(n, kk):
                            d = n.display
                            dp = nd2.display
                            for kkk in dp.__dict__.keys():
                                if hasattr(d, kkk):
                                    exec('dp.' + kkk + '=d.' + kkk)

                            nd2.display = dp

                    elif kk == 'apc':
                        if hasattr(n, kk):
                            ab = n.apc
                            ac = nd2.apc
                            for kkk in ac.__dict__.keys():
                                if hasattr(ab, kkk):
                                    exec('ac.' + kkk + '=ab.' + kkk)

                            nd2.apc = ac

                    elif kk == 'hsqc':
                        if hasattr(n, kk):
                            h = n.hsqc
                            hp = nd2.hsqc
                            for kkk in hp.__dict__.keys():
                                if hasattr(h, kkk):
                                    # if kkk != 'hsqc_data':
                                    exec('hp.' + kkk + '=h.' + kkk)
                                    #
                                    # else:
                                    #    hd = h.hsqc_data
                                    #    hdp = hp.hsqc_data
                                    #    for kkkk in hdp.__dict__.keys():
                                    #        if hasattr(hd, kkkk):
                                    #            exec('hdp.' + kkkk + '=hd.' + kkkk)
                                    #
                                    #    hp.hsqc_data = hdp

                            nd2.hsqc = hp

                if vver == '0.1':
                    nd2.ver = '0.1'

                nda[k].append(nd2)
                nd2 = []
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'titleFile.txt')
                f = open(f_name, 'r')
                nda[k][l].title = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'acqusText.txt')
                f = open(f_name, 'r')
                nda[k][l].acq.acqus_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'acqu2sText.txt')
                f = open(f_name, 'r')
                nda[k][l].acq.acqu2s_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'acqu3sText.txt')
                f = open(f_name, 'r')
                nda[k][l].acq.acqu3s_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'procsText.txt')
                f = open(f_name, 'r')
                nda[k][l].proc.procs_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'proc2sText.txt')
                f = open(f_name, 'r')
                nda[k][l].proc.proc2s_text = f.read()
                f.close()
                f_name = os.path.join(os.path.join(os.path.join(data_set_name, data_sets[k]), data_set_exps[k][l]),
                                      'proc3sText.txt')
                f = open(f_name, 'r')
                nda[k][l].proc.proc3s_text = f.read()
                f.close()

        #self.cf.current_directory = data_set_name
        #self.cf.save_config()
        #if len(self.nmrdat) - 1 < self.s:
        #    self.s = 0
        #
        #if len(self.nmrdat[self.s]) - 1< self.e:
        #    self.e = 0

        return nda
        # end load2

    def load_mat(self, file_name=''):
        if len(file_name) == 0:
            return

        m = mat73.loadmat(file_name)
        if 'NMRDAT' not in m.keys():
            return

        self.clear()
        self.s = 0
        self.e = 0
        for exp in range(len(m['NMRDAT']['ACQUSText'])):
            gc.collect()
            nd1 = nd.NmrData()
            acqus = m['NMRDAT']['ACQUSText'][exp][0]
            nd1.acq.acqus_text += acqus[0][0]
            counter = -1
            for idx in range(1, len(acqus)):
                if acqus[idx][0].find('##') > -1:
                    counter = -1
                    nd1.acq.acqus_text += '\n'
                    nd1.acq.acqus_text += acqus[idx][0]
                elif acqus[idx][0].find('(') > -1 and acqus[idx][0].find('(') > -1:
                    counter = 0
                    nd1.acq.acqus_text += ' '
                    nd1.acq.acqus_text += acqus[idx][0]
                    nd1.acq.acqus_text += '\n'
                else:
                    if counter > -1:
                        counter += 1
                    nd1.acq.acqus_text += ' '
                    nd1.acq.acqus_text += acqus[idx][0]
                    if counter > 32:
                        nd1.acq.acqus_text += '\n'
                        counter = 0

            try:
                if hasattr(m['NMRDAT']['ACQU2SText'], 'keys'):
                    acqu2s = m['NMRDAT']['ACQU2SText'][exp][0]
                    nd1.acq.acqu2s_text += acqu2s[0][0]
                    for idx in range(1, len(acqu2s)):
                        if acqu2s[idx][0].find('##') > -1:
                            nd1.acq.acqu2s_text += '\n'
                        else:
                            nd1.acq.acqu2s_text += ' '

                        nd1.acq.acqu2s_text += acqu2s[idx][0]
            except:
                pass

            try:
                if hasattr(m['NMRDAT']['ACQU3SText'], 'keys'):
                    acqu3s = m['NMRDAT']['ACQU3SText'][exp][0]
                    nd1.acq.acqu3s_text += acqu3s[0][0]
                    for idx in range(1, len(acqu3s)):
                        if acqu3s[idx][0].find('##') > -1:
                            nd1.acq.acqu3s_text += '\n'
                        else:
                            nd1.acq.acqu3s_text += ' '

                        nd1.acq.acqu3s_text += acqu3s[idx][0]
            except:
                pass

            nd1.acq.parse_reg_ex()
            if nd1.acq.group_delay == 0.0:
                if copy(m['NMRDAT']['ACQUS'][exp]['GRPDLY'][0]) + 0 == 0:
                    nd1.acq.set_group_delay()
                else:
                    nd1.acq.group_delay = copy(m['NMRDAT']['ACQUS'][exp]['GRPDLY'][0]) + 0

            nd1.acq.sfo2 = nd1.acq.bf2 + nd1.acq.o2 / 1000000.0

            try:
                procs = m['NMRDAT']['PROCSText'][exp][0]
                nd1.proc.procs_text += procs[0][0]
                for idx in range(1, len(procs)):
                    if procs[idx][0].find('##') > -1:
                        nd1.proc.procs_text += '\n'
                    else:
                        nd1.proc.procs_text += ' '

                    nd1.proc.procs_text += procs[idx][0]
            except:
                pass

            try:
                if hasattr(m['NMRDAT']['PROC2SText'], 'keys'):
                    proc2s = m['NMRDAT']['PROC2SText'][exp][0]
                    nd1.proc.proc2s_text += proc2s[0][0]
                    for idx in range(1, len(proc2s)):
                        if proc2s[idx][0].find('##') > -1:
                            nd1.proc.proc2s_text += '\n'
                        else:
                            nd1.proc.proc2s_text += ' '

                        nd1.proc.proc2s_text += proc2s[idx][0]
            except:
                pass

            try:
                if hasattr(m['NMRDAT']['PROC3SText'], 'keys'):
                    proc3s = m['NMRDAT']['PROC3SText'][exp][0]
                    nd1.proc.proc3s_text += proc3s[0][0]
                    for idx in range(1, len(proc3s)):
                        if proc3s[idx][0].find('##') > -1:
                            nd1.proc.proc3s_text += '\n'
                        else:
                            nd1.proc.proc3s_text += ' '

                        nd1.proc.proc3s_text += proc3s[idx][0]
            except:
                pass
            try:
                nd1.proc.parse_reg_ex()
            except:
                pass

            if nd1.proc.axis_nucleus[1] != 'off':
                nd1.display.y_label = nd1.proc.axis_nucleus[1]

            nd1.orig_data_set = m['NMRDAT']['NAME'][exp]
            nd1.title = copy(m['NMRDAT']['COMMENT'][exp])
            nd1.title_orig = copy(m['NMRDAT']['COMMENT'][exp])
            if m['NMRDAT']['ACQUS'][exp]['DIM'][0] == 1:
                npts1 = 1
                npts2 = len(m['NMRDAT']['MAT'][exp])
                fidpts1 = 1
                fidpts2 = len(m['NMRDAT']['SER'][exp])
                nd1.acq.n_data_points[0] = fidpts2
                nd1.fid = np.resize(nd1.fid, (fidpts1, fidpts2))
                nd1.spc = np.resize(nd1.spc, (npts1, npts2))
                nd1.fid[0] = m['NMRDAT']['SER'][exp]
                nd1.spc[0] = m['NMRDAT']['MAT'][exp]
                nd1.dim = 1
                nd1.ref_shift[0] = copy(m['NMRDAT']['PROC'][exp]['REF'][0][0])
                nd1.ref_point[0] = npts2 - copy(m['NMRDAT']['PROC'][exp]['REF'][0][1])
                # deref ???
                # https://stackoverflow.com/questions/15011674/is-it-possible-to-dereference-variable-ids
                nd1.proc.n_points[0] = npts2
                nd1.proc.ph0[0] = (-float(m['NMRDAT']['PROC'][exp]['PH0'][0]) - 90.0) % 360.0
                nd1.proc.ph1[0] = -float(m['NMRDAT']['PROC'][exp]['PH1'][0])
                nd1.proc.lb[0] = float(m['NMRDAT']['PROC'][exp]['LB'][0])
                nd1.proc.gb[0] = float(m['NMRDAT']['PROC'][exp]['GB'][0])
                nd1.proc.ssb[0] = float(m['NMRDAT']['PROC'][exp]['SSB'][0])
            else:
                npts1 = len(m['NMRDAT']['MAT'][exp])
                npts2 = len(m['NMRDAT']['MAT'][exp][0])
                fidpts1 = len(m['NMRDAT']['SER'][exp])
                fidpts2 = len(m['NMRDAT']['SER'][exp][0])
                nd1.acq.n_data_points[0] = fidpts2
                nd1.acq.n_data_points[1] = fidpts1
                nd1.fid = np.resize(nd1.fid, (fidpts1, fidpts2))
                nd1.spc = np.resize(nd1.spc, (npts1, npts2))
                nd1.fid = m['NMRDAT']['SER'][exp]
                nd1.spc = m['NMRDAT']['MAT'][exp]
                nd1.dim = 2
                nd1.ref_shift[0] = copy(m['NMRDAT']['PROC'][exp]['REF'][0][0])
                nd1.ref_point[0] = npts2 - np.copy(m['NMRDAT']['PROC'][exp]['REF'][0][1])
                nd1.ref_shift[1] = copy(m['NMRDAT']['PROC'][exp]['REF'][1][0])
                nd1.ref_point[1] = npts1 - np.copy(m['NMRDAT']['PROC'][exp]['REF'][1][1])
                nd1.proc.n_points[0] = npts2
                nd1.proc.n_points[1] = npts1
                nd1.acq.sw_h[0] = float(m['NMRDAT']['PROC'][exp]['REF'][0][3])
                nd1.acq.sw_h[1] = float(m['NMRDAT']['ACQUS'][exp]['SW_h'][1])
                nd1.proc.ph0[0] = (-float(m['NMRDAT']['PROC'][exp]['PH0'][0]) - 90.0) % 360.0
                nd1.proc.ph1[0] = -float(m['NMRDAT']['PROC'][exp]['PH1'][0])
                nd1.proc.ph0[1] = -float(m['NMRDAT']['PROC'][exp]['PH0'][1])
                nd1.proc.ph1[1] = -float(m['NMRDAT']['PROC'][exp]['PH1'][1])
                nd1.proc.lb[0] = float(m['NMRDAT']['PROC'][exp]['LB'][0])
                nd1.proc.gb[0] = float(m['NMRDAT']['PROC'][exp]['GB'][0])
                nd1.proc.ssb[0] = float(m['NMRDAT']['PROC'][exp]['SSB'][0])
                nd1.proc.lb[1] = float(m['NMRDAT']['PROC'][exp]['LB'][1])
                nd1.proc.gb[1] = float(m['NMRDAT']['PROC'][exp]['GB'][1])
                nd1.proc.ssb[1] = float(m['NMRDAT']['PROC'][exp]['SSB'][1])

            nd1.proc.sw_h = np.copy(nd1.acq.sw_h)
            nd1.acq.manufacturer = 'Bruker'
            nd1.calc_ppm()
            if hasattr(m['metaboSpc'], 'keys'):
                if type(m['metaboSpc']['baseline']['spline']['baseline_points']) == np.ndarray:
                    nd1.spline_baseline.linear_spline = int(m['metaboSpc']['baseline']['spline']['linear_points'])
                    nd1.spline_baseline.average_points = int(
                        m['metaboSpc']['baseline']['spline']['baseline_average_points'])
                    nd1.spline_baseline.baseline_points = nd1.points2ppm(
                        npts2 - np.array(m['metaboSpc']['baseline']['spline']['baseline_points'], dtype=int))
                    nd1.add_baseline_points()

            self.nmrdat[self.s].append(nd1)

        if hasattr(m['metaboSpc'], 'keys'):
            if type(m['metaboSpc']['exclude']['start']) == np.ndarray:
                try:
                    self.pp.exclude_start = np.sort(np.round(1e4 * np.array(m['metaboSpc']['exclude']['stop'])) / 1e4)
                    self.pp.exclude_end = np.sort(np.round(1e4 * np.array(m['metaboSpc']['exclude']['start'])) / 1e4)
                except:
                    self.pp.exclude_start = np.round(1e4 * np.array(m['metaboSpc']['exclude']['stop'])) / 1e4
                    self.pp.exclude_end = np.round(1e4 * np.array(m['metaboSpc']['exclude']['start'])) / 1e4

            if type(m['metaboSpc']['compress1']['start']) == np.ndarray:
                try:
                    self.pp.seg_start = np.sort(np.round(1e4 * np.array(m['metaboSpc']['compress1']['stop'])) / 1e4)
                    self.pp.seg_end = np.sort(np.round(1e4 * np.array(m['metaboSpc']['compress1']['start'])) / 1e4)
                except:
                    self.pp.seg_start = np.round(1e4 * np.array([m['metaboSpc']['compress1']['stop']])) / 1e4
                    self.pp.seg_end = np.round(1e4 * np.array([m['metaboSpc']['compress1']['start']])) / 1e4


            if 'noiseFilter' in m['metaboSpc'].keys():
                if type(m['metaboSpc']['noiseFilter']['threshold']) == np.ndarray:
                    self.pp.noise_threshold = float(m['metaboSpc']['noiseFilter']['threshold'])
                    self.pp.noise_start = float(m['metaboSpc']['noiseFilter']['regEnd'])
                    self.pp.noise_end = float(m['metaboSpc']['noiseFilter']['regStart'])

            if type(m['metaboSpc']['bucket']['bucketSizePoints']) == np.ndarray:
                self.pp.bucket_points = float(m['metaboSpc']['bucket']['bucketSizePoints'])
                ppm_per_point = abs(nd1.ppm1[0] - nd1.ppm1[1])
                bucket_points = self.pp.bucket_points
                bucket_ppm = np.round(1e4 * bucket_points * ppm_per_point) / 1e4
                self.pp.bucket_points = bucket_points
                self.pp.bucket_ppm = bucket_ppm

            if type(m['metaboSpc']['compress2']['start']) == np.ndarray:
                try:
                    self.pp.compress_start = np.sort(
                        np.round(1e4 * np.array(m['metaboSpc']['compress2']['stop'])) / 1e4)
                    self.pp.compress_end = np.sort(np.round(1e4 * np.array(m['metaboSpc']['compress2']['start'])) / 1e4)
                except:
                    self.pp.compress_start = np.round(1e4 * np.array([m['metaboSpc']['compress2']['stop']])) / 1e4
                    self.pp.compress_end = np.round(1e4 * np.array([m['metaboSpc']['compress2']['start']])) / 1e4

            if type(m['metaboSpc']['scale2']['pqn']) == np.ndarray:
                if float(m['metaboSpc']['scale2']['pqn']) == 1.0:
                    self.pp.scale_pqn = True
                else:
                    self.pp.scale_pqn = False

            self.pp.auto_scaling = bool(m['metaboSpc']['advancedScale']['use_autoscale'])
            self.pp.pareto_scaling = bool(m['metaboSpc']['advancedScale']['use_paretoscale'])
            self.pp.g_log_transform = bool(m['metaboSpc']['advancedScale']['use_glogscale'])
            self.pp.var_lambda = float(m['metaboSpc']['advancedScale']['lambda'])
            self.pp.var_y0 = float(m['metaboSpc']['advancedScale']['glog_minlev'])
            self.reset_data_pre_processing()
        # end load_mat

    def noise_filtering(self):
        val = self.pp.noise_threshold * self.pp.std_val
        for k in range(len(self.nmrdat[self.s])):
            idx = np.where(self.nmrdat[self.s][k].spc[0].real < val)
            self.deselect2[idx] += np.ones(len(idx))
            # if False:
            #    idx2 = np.where(self.nmrdat[self.s][k].spc[0].real < 0)
            #    self.nmrdat[self.s][k].spc[0][idx2] = np.zeros(len(idx2))

        # end noise_filtering

    def noise_filtering_init(self):
        spc_idx = np.where(
            (self.nmrdat[self.s][0].ppm1 > self.pp.noise_start) & (self.nmrdat[self.s][0].ppm1 < self.pp.noise_end))
        self.pp.std_val = np.std(self.nmrdat[self.s][0].spc[0][spc_idx].real)
        # end noise_filtering_init

    def pjres(self, set=-1, mode='skyline'):
        if set < 1:
            set = self.s + 2

        if (len(self.nmrdat) < set):
            self.nmrdat.append([])

        self.nmrdat[set - 1] = []
        for k in range(len(self.nmrdat[self.s])):
            nd1 = nd.NmrData()
            self.nmrdat[set - 1].append(nd1)
            self.nmrdat[set - 1][k].dim = 1
            self.nmrdat[set - 1][k].fid = [[]]
            self.nmrdat[set - 1][k].acq = self.nmrdat[self.s][k].acq
            self.nmrdat[set - 1][k].proc = self.nmrdat[self.s][k].proc
            self.nmrdat[set - 1][k].display = self.nmrdat[self.s][k].display
            self.nmrdat[set - 1][k].ppm1 = np.resize(self.nmrdat[set - 1][k].ppm1, (len(self.nmrdat[self.s][k].ppm1)))
            self.nmrdat[set - 1][k].ppm1 = np.copy(self.nmrdat[self.s][k].ppm1)
            self.nmrdat[set - 1][k].spc = np.resize(self.nmrdat[set - 1][k].spc,
                                                    (1, len(self.nmrdat[self.s][k].spc[0])))
            self.nmrdat[set - 1][k].ref_shift = copy(self.nmrdat[self.s][k].ref_shift)
            self.nmrdat[set - 1][k].ref_point = self.nmrdat[self.s][k].ref_point
            self.nmrdat[set - 1][k].title = "pJres spectrum\n" + self.nmrdat[self.s][k].title
            self.nmrdat[set - 1][k].pjres_mode = mode
            self.nmrdat[set - 1][k].projected_j_res = True
            self.nmrdat[set - 1][k].orig_j_res_set = self.s
            self.nmrdat[set - 1][k].orig_j_res_exp = k
            if mode == 'skyline':
                self.nmrdat[set - 1][k].spc[0] = np.max(self.nmrdat[self.s][k].spc, 0)
            else:
                self.nmrdat[set - 1][k].spc[0] = np.sum(self.nmrdat[self.s][k].spc, 0)

        # end jresProject

    def plot_spc(self):  # pragma: no cover
        pl.plot()
        ax = pl.gca()
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        ax.clear()
        if len(self.nmrdat[self.s]) == 0:
            return

        if len(self.nmrdat[self.s][self.e].spc) == 0:
            return

        d = self.nmrdat[self.s][self.e].display
        if d.pos_col == "RGB":
            pos_col = d.pos_col_rgb
        else:
            pos_col = d.pos_col

        if d.neg_col == "RGB":
            neg_col = d.neg_col_rgb
        else:
            neg_col = d.neg_col

        pos_col = matplotlib.colors.to_hex(pos_col)
        neg_col = matplotlib.colors.to_hex(neg_col)
        xlabel = d.x_label + " [" + d.axis_type1 + "]"
        ylabel = d.y_label + " [" + d.axis_type2 + "]"
        if self.nmrdat[self.s][self.e].dim == 1:
            for k in range(len(self.nmrdat[self.s])):
                if k != self.e and self.nmrdat[self.s][k].display.display_spc == True:
                    d = self.nmrdat[self.s][k].display
                    if d.pos_col == "RGB":
                        pos_col = d.pos_col_rgb
                    else:
                        pos_col = d.pos_col

                    if d.neg_col == "RGB":
                        neg_col = d.neg_col_rgb
                    else:
                        neg_col = d.neg_col

                    pos_col = matplotlib.colors.to_hex(pos_col)
                    neg_col = matplotlib.colors.to_hex(neg_col)
                    pl.plot(self.nmrdat[self.s][k].ppm1, self.nmrdat[self.s][k].spc[0].real, color=pos_col)

            d = self.nmrdat[self.s][self.e].display
            if d.pos_col == "RGB":
                pos_col = d.pos_col_rgb
            else:
                pos_col = d.pos_col

            if d.neg_col == "RGB":
                neg_col = d.neg_col_rgb
            else:
                neg_col = d.neg_col

            pos_col = matplotlib.colors.to_hex(pos_col)
            neg_col = matplotlib.colors.to_hex(neg_col)
            xlabel = d.x_label + " [" + d.axis_type1 + "]"
            ylabel = d.y_label + " [" + d.axis_type2 + "]"
            pl.plot(self.nmrdat[self.s][self.e].ppm1, self.nmrdat[self.s][self.e].spc[0].real, color=pos_col)
            ax = pl.gca()
            ax.set_xlabel(xlabel)
            ax.autoscale()
            ax.invert_xaxis()
            if self.keep_zoom == True:
                if xlim[0] != -0.05 and xlim[1] != 1.05 and ylim[0] != -0.05 and ylim[1] != 1.05:
                    ax.set_xlim(xlim)
                    ax.set_ylim(ylim)

        else:
            mm = np.max(np.abs(self.nmrdat[self.s][self.e].spc.real))
            pos_lev = np.linspace(d.min_level * mm, d.max_level * mm, d.n_levels)
            neg_lev = np.linspace(-d.max_level * mm, -d.min_level * mm, d.n_levels)
            pl.contour(self.nmrdat[self.s][self.e].ppm1,
                       self.nmrdat[self.s][self.e].ppm2,
                       self.nmrdat[self.s][self.e].spc.real, pos_lev, colors=pos_col,
                       linestyles='solid', antialiased=True)
            pl.contour(self.nmrdat[self.s][self.e].ppm1,
                       self.nmrdat[self.s][self.e].ppm2,
                       self.nmrdat[self.s][self.e].spc.real, neg_lev, colors=neg_col,
                       linestyles='solid', antialiased=True)
            ax = pl.gca()
            ax.set_xlabel(xlabel)
            ax.set_ylabel(ylabel)
            ax.autoscale()
            ax.invert_xaxis()
            ax.invert_yaxis()

        if self.cf.mode == 'dark':
            bg = (50 / 255, 58 / 255, 72 / 255)
            fg = (255 / 255, 255 / 255, 255 / 255)
            pl.gcf().set_facecolor(bg)
            pl.gca().set_facecolor(bg)
            pl.gca().xaxis.label.set_color(fg)
            pl.gca().yaxis.label.set_color(fg)
            pl.gca().tick_params(axis='x', colors=fg)
            pl.gca().tick_params(axis='y', colors=fg)
            pl.gca().spines['bottom'].set_color(fg)
            pl.gca().spines['top'].set_color(fg)
            pl.gca().spines['left'].set_color(fg)
            pl.gca().spines['right'].set_color(fg)

        # end plot_spc

    def pre_proc_init(self):
        self.pp.init(len(self.nmrdat[self.s]))
        # end pre_proc_init

    def read_spc(self, data_set_name='', data_set_number='', dataset=1):
        if len(data_set_name) == 0:
            data_set_name = self.data_set_name
            if len(data_set_name) == 0:
                return

        if len(data_set_number) == 0:
            data_set_number = self.data_set_number
            if len(data_set_number) == 0:
                return

        if len(self.nmrdat) < dataset:
            self.nmrdat.append([])

        self.s = dataset - 1
        self.e = len(self.nmrdat[self.s])
        nd1 = nd.NmrData()
        nd1.data_set_name = data_set_name
        nd1.data_set_number = data_set_number
        nd1.read_spc()
        if nd1.acq.pul_prog_name.find('stddiff') > 0:
            nd1.fid[0] -= nd1.fid[1]
            nd1.fid = np.copy(np.delete(nd1.fid, 1, 0))
            nd1.dim = 1

        self.nmrdat[self.s].append(nd1)
        # end read_spc

    def read_spcs(self, data_path, data_exp, dataset=1):
        if len(data_exp) > 1:
            for k in range(len(data_exp)):
                self.read_spc(data_path[0], str(data_exp[k]), dataset)

        else:
            for k in range(len(data_path)):
                self.read_spc(data_path[k], str(data_exp[0]), dataset)

    # end read_spcs

    def read_title_file_information_excel(self, file_name=''):
        if len(file_name) == 0:
            return

        xls = pd.read_excel(file_name).fillna('')
        return xls

    def read_nmrpipe_spc(self, data_set_name, data_set_number, proc_data_name='test.dat', ft_dir=''):
        self.e = len(self.nmrdat[self.s])
        nd1 = nd.NmrData()
        nd1.data_set_name = data_set_name
        nd1.data_set_number = data_set_number
        nd1.read_spc()
        nd1.read_pipe_2d(data_set_name + os.sep + data_set_number + '.proc', proc_data_name, ft_dir)
        nd1.acq.sw[0] = nd1.acq.sw[0] * len(nd1.spc[0]) / 2 ** math.ceil(math.log(len(nd1.spc[0]), 2))
        nd1.acq.sw_h[0] = nd1.acq.sw_h[0] * len(nd1.spc[0]) / 2 ** math.ceil(math.log(len(nd1.spc[0]), 2))
        nd1.calc_ppm()
        self.nmrdat[self.s].append(nd1)
        # end read_spc

    def read_nmrpipe_spcs(self, data_path, data_exp, proc_data_name='test.dat', ft_dir=''):
        if len(data_exp) > 1:
            for k in range(len(data_exp)):
                self.read_nmrpipe_spc(data_path[0], str(data_exp[k]), proc_data_name, ft_dir)

        else:
            for k in range(len(data_path)):
                self.read_nmrpipe_spc(data_path[k], str(data_exp[0]), proc_data_name, ft_dir)

    # end read_spcs

    def reference1d_all(self, old_ppm=0.0, new_ppm=0.0, find_maximum=True):
        s = self.s
        e = self.e
        for k in range(len(self.nmrdat[s])):
            if self.nmrdat[s][k].dim != 1:
                break

            ref_point = len(self.nmrdat[s][k].spc[0]) - self.nmrdat[s][k].ppm2points(old_ppm) - 1
            diff_pts = np.linspace(-1, 1, 3, dtype=int)
            ref_pts = ref_point * np.ones(3, dtype=int) + diff_pts
            found_maximum = False
            while not found_maximum:
                if find_maximum:
                    max_idx = \
                    np.where(self.nmrdat[s][k].spc[0][ref_pts] == np.max(self.nmrdat[s][k].spc[0][ref_pts]))[0][0]
                else:
                    max_idx = \
                    np.where(self.nmrdat[s][k].spc[0][ref_pts] == np.min(self.nmrdat[s][k].spc[0][ref_pts]))[0][0]

                ref_point += diff_pts[max_idx]
                ref_pts = ref_point * np.ones(3, dtype=int) + diff_pts
                if max_idx == 1:
                    found_maximum = True

            self.nmrdat[s][k].ref_point[0] = len(self.nmrdat[s][k].spc[0]) - ref_point - 1
            self.nmrdat[s][k].ref_shift[0] = copy(new_ppm)
            self.nmrdat[s][k].ref = 'manual'
            self.nmrdat[s][k].calc_ppm()
        # end reference1d_all

    def reset_data_pre_processing(self, autoref=True):
        if not self.nmrdat[self.s][0].projected_j_res:
            self.ft_all()
            if self.nmrdat[self.s][0].proc.autobaseline:
                self.autobaseline1d_all()
            else:
                self.baseline1d_all()

            if autoref:
                self.auto_ref_all()

            self.shift_ref()
            if len(self.nmrdat[self.s][self.e].spline_baseline.baseline_points) > 0:
                for k in range(len(self.nmrdat[self.s])):
                    self.nmrdat[self.s][k].spline_baseline = self.nmrdat[self.s][self.e].spline_baseline
                    self.nmrdat[self.s][k].add_baseline_points()
                    self.nmrdat[self.s][k].corr_spline_baseline()

        else:
            s = self.s
            e = self.e
            self.s = self.nmrdat[s][e].orig_j_res_set
            self.e = self.nmrdat[s][e].orig_j_res_exp
            self.pjres(s + 1, self.nmrdat[s][e].pjres_mode)
            self.s = s
            self.e = e

        # end reset_data_pre_processing

    def reshape_title(self, n_rows=2):
        return_text = self.nmrdat[self.s][self.e].reshape_title(n_rows)
        return return_text
        # end reshape_title

    def reshape_titles(self, n_rows=2):
        all_fine = True
        for k in range(len(self.nmrdat[self.s])):
            if k == self.e:
                print_text = True
            else:
                print_text = False

            return_text = self.nmrdat[self.s][k].reshape_title(n_rows=n_rows, print_text=print_text, exp=k)
            if return_text != 'Succesfully reshaped title':
                all_fine = False

        return all_fine

    def save(self, data_set_name):
        if len(data_set_name) == 0:
            return

        try:
            os.makedirs(data_set_name)
        except:
            pass

        f_name = os.path.join(data_set_name, 'curPars.dat')
        f = open(f_name, 'wb')
        pickle.dump([self.file_format_version, self.s, self.e, self.pp, self.deselect, self.deselect2, self.cmd_buffer,
                     self.cmd_idx, self.script, self.console, self.tmsp_conc], f)
        f.close()
        for k in range(len(self.nmrdat)):
            setPath = os.path.join(data_set_name, str(k + 1))
            try:
                os.makedirs(setPath)
            except:
                pass

            for l in range(len(self.nmrdat[k])):
                exp_path = os.path.join(setPath, str(l + 1))
                try:
                    os.makedirs(exp_path)
                except:
                    pass

                f_name = os.path.join(exp_path, 'nmrDataSet.dat')
                f = open(f_name, 'wb')
                self.nmrdat[k][l].ver = self.ver
                pickle.dump(self.nmrdat[k][l], f)
                f.close()
                f_name = os.path.join(exp_path, 'titleFile.txt')
                f = open(f_name, 'w')
                f.write(self.nmrdat[k][l].title)
                f.close()
                f_name = os.path.join(exp_path, 'procsText.txt')
                f = open(f_name, 'w')
                f.write(self.nmrdat[k][l].proc.procs_text)
                f.close()
                f_name = os.path.join(exp_path, 'proc2sText.txt')
                f = open(f_name, 'w')
                f.write(self.nmrdat[k][l].proc.proc2s_text)
                f.close()
                f_name = os.path.join(exp_path, 'proc3sText.txt')
                f = open(f_name, 'w')
                f.write(self.nmrdat[k][l].proc.proc3s_text)
                f.close()
                f_name = os.path.join(exp_path, 'acqusText.txt')
                f = open(f_name, 'w')
                f.write(self.nmrdat[k][l].acq.acqus_text)
                f.close()
                f_name = os.path.join(exp_path, 'acqu2sText.txt')
                f = open(f_name, 'w')
                f.write(self.nmrdat[k][l].acq.acqu2s_text)
                f.close()
                f_name = os.path.join(exp_path, 'acqu3sText.txt')
                f = open(f_name, 'w')
                f.write(self.nmrdat[k][l].acq.acqu3s_text)
                f.close()

        self.cf.current_directory = data_set_name
        self.cf.save_config()
        # end save

    def scale_spectra(self):
        n_spc = len(self.nmrdat[self.s])
        npts = len(self.nmrdat[self.s][0].spc[0])
        if self.pp.scale_spectra_ref_spc > 0:
            ref_spc = self.nmrdat[self.s][self.pp.scale_spectra_ref_spc - 1].spc[0].real
        else:
            ref_spcs = np.zeros((n_spc, npts))
            for k in range(n_spc):
                ref_spcs[k] = self.nmrdat[self.s][k].spc[0].real

            if self.pp.seg_align_ref_spc == 0:
                ref_spc = np.mean(ref_spcs, 0)
            else:
                ref_spc = np.median(ref_spcs, 0)

            ref_spcs = np.array([[]])

        scale = np.ones(n_spc)
        if self.pp.scale_pqn is True:
            for k in range(n_spc):
                scale_vect = self.nmrdat[self.s][k].spc[0][np.where(ref_spc != 0)].real / ref_spc[
                    np.where(ref_spc != 0)]
                self.nmrdat[self.s][k].spc[0] /= np.median(scale_vect[np.where(scale_vect != 0)])
                self.pp.spc_scale[k] = np.median(scale_vect[np.where(scale_vect != 0)])
        else:
            if self.pp.preserve_overall_scale is True:
                for k in range(n_spc):
                    scale[k] = np.sum(self.nmrdat[self.s][k].spc[0]).real

            for k in range(n_spc):
                spc_sum = np.sum(self.nmrdat[self.s][k].spc[0]).real
                self.nmrdat[self.s][k].spc[0] /= spc_sum
                if self.pp.scale_spectra_ref_spc > 0:
                    self.nmrdat[self.s][k].spc[0] *= scale[self.pp.scale_spectra_ref_spc - 1]
                    self.pp.spc_scale[k] = spc_sum / scale[self.pp.scale_spectra_ref_spc - 1]
                else:
                    self.nmrdat[self.s][k].spc[0] *= np.max(scale)
                    self.pp.spc_scale[k] = spc_sum / np.max(scale)

        # end scale_spectra

    def segmental_alignment(self):
        if self.pp.seg_align_ref_spc > 0:
            kk = self.pp.seg_align_ref_spc - 1
        else:
            kk = 0
        seg_start = self.nmrdat[self.s][kk].ppm2points(self.pp.seg_start, 0)
        seg_end = self.nmrdat[self.s][kk].ppm2points(self.pp.seg_end, 0)
        npts = len(self.nmrdat[self.s][kk].spc[0])
        n_spc = len(self.nmrdat[self.s])
        exclude_start = np.zeros((len(seg_start), n_spc))
        exclude_end = np.zeros((len(seg_start), n_spc))
        if self.pp.seg_align_ref_spc > 0:
            ref_spc = self.nmrdat[self.s][self.pp.seg_align_ref_spc - 1].spc[0]
        else:
            ref_spcs = np.zeros((n_spc, npts))
            for k in range(n_spc):
                ref_spcs[k] = self.nmrdat[self.s][k].spc[0].real

            if self.pp.seg_align_ref_spc == 0:
                ref_spc = np.mean(ref_spcs, 0)
            else:
                ref_spc = np.median(ref_spcs, 0)

            ref_spcs = np.array([[]])

        pos_shift = np.zeros((n_spc, len(seg_start)))
        neg_shift = np.zeros((n_spc, len(seg_start)))
        for k in range(n_spc):
            self.nmrdat[self.s][k].seg_align_corr_before = []
            self.nmrdat[self.s][k].seg_align_corr_after = []
            self.nmrdat[self.s][k].seg_align_corr_ppm1 = []
            self.nmrdat[self.s][k].seg_align_corr_ppm2 = []
            if k != self.pp.seg_align_ref_spc - 1:
                for l in range(len(seg_start)):
                    start_pts = npts - seg_end[l]
                    end_pts = npts - seg_start[l]
                    corr_spc1 = np.copy(ref_spc[start_pts:end_pts].real)
                    corr_spc2 = self.nmrdat[self.s][k].spc[0][start_pts:end_pts].real
                    max_shift = len(corr_spc1) - 1
                    zeros = np.zeros(len(corr_spc1))
                    corr_spc1 = np.append(np.append(zeros, corr_spc1), zeros)
                    corr_spc2 = np.append(np.append(zeros, corr_spc2), zeros)
                    shifts = np.linspace(-max_shift, max_shift, 2 * max_shift + 1, dtype='int')
                    corr_vect = np.zeros(2 * max_shift + 1)
                    spc_shift = 0
                    self.nmrdat[self.s][k].seg_align_corr_before.append(float(np.corrcoef(corr_spc1, corr_spc2)[0][1]))
                    for m in shifts:
                        corr_vect[m + max_shift] = np.corrcoef(corr_spc1, np.roll(corr_spc2, m))[0][1]

                    self.nmrdat[self.s][k].seg_align_corr_after.append(float(corr_vect[np.where(corr_vect == np.max(corr_vect))[0][0]]))
                    self.nmrdat[self.s][k].seg_align_corr_ppm1.append(float(self.pp.seg_start[l]))
                    self.nmrdat[self.s][k].seg_align_corr_ppm2.append(float(self.pp.seg_end[l]))
                    max_corr_shifts = shifts[np.where(corr_vect == np.max(corr_vect))]
                    min_shift = np.where(np.abs(max_corr_shifts) == np.min(np.abs(max_corr_shifts)))
                    if np.max(corr_vect) > 0.8:
                        spc_shift = max_corr_shifts[min_shift][0]
                        corr_spc2 = self.nmrdat[self.s][k].spc[0][start_pts:end_pts]
                        corr_spc2 = np.append(np.append(zeros, corr_spc2), zeros)
                        corr_spc2 = np.roll(corr_spc2, spc_shift)[max_shift + 1:2 * max_shift + 2]
                        self.nmrdat[self.s][k].spc[0][start_pts:end_pts] = np.copy(corr_spc2)
                        ex_sta = 0
                        ex_end = 0
                        if spc_shift < 0:
                            neg_shift[k][l] = 0 - spc_shift
                            ex_end = self.nmrdat[self.s][self.pp.seg_align_ref_spc - 1].points2ppm(npts - end_pts -
                                                                                                   spc_shift, 0)
                            ex_sta = self.nmrdat[self.s][self.pp.seg_align_ref_spc - 1].points2ppm(npts - end_pts, 0)
                        elif spc_shift > 0:
                            pos_shift[k][l] = spc_shift
                            ex_end = self.nmrdat[self.s][self.pp.seg_align_ref_spc - 1].points2ppm(npts - start_pts, 0)
                            ex_sta = self.nmrdat[self.s][self.pp.seg_align_ref_spc - 1].points2ppm(
                                npts - start_pts - spc_shift, 0)

                        exclude_start[l][k] = ex_sta
                        exclude_end[l][k] = ex_end
            else:
                for l in range(len(seg_start)):
                    self.nmrdat[self.s][k].seg_align_corr_before.append(1.0)
                    self.nmrdat[self.s][k].seg_align_corr_after.append(1.0)
                    self.nmrdat[self.s][k].seg_align_corr_ppm1.append(float(self.pp.seg_start[l]))
                    self.nmrdat[self.s][k].seg_align_corr_ppm2.append(float(self.pp.seg_end[l]))


        ps = np.max(pos_shift, 0)
        ns = np.max(neg_shift, 0)
        ps2 = np.transpose(pos_shift)
        ns2 = np.transpose(neg_shift)
        exclude_start2 = np.array([])
        exclude_end2 = np.array([])
        for k in range(len(pos_shift[0])):
            l = np.where(ps2[k] == ps[k])[0][0]
            if ps[k] > 0:
                s_val = math.floor(1e4 * exclude_start[k][l]) / 1e4
                e_val = math.floor(1e4 * exclude_end[k][l]) / 1e4
                if s_val not in self.pp.exclude_start and e_val not in self.pp.exclude_end:
                    exclude_start2 = np.append(exclude_start2, math.floor(1e4 * s_val) / 1e4)
                    exclude_end2 = np.append(exclude_end2, math.floor(1e4 * e_val) / 1e4)
                    # self.pp.exclude_start = np.append(self.pp.exclude_start, s_val)
                    # self.pp.exclude_end   = np.append(self.pp.exclude_end,   e_val)

            l = np.where(ns2[k] == ns[k])[0][0]
            if ns[k] > 0:
                s_val = math.floor(1e4 * exclude_start[k][l]) / 1e4
                e_val = math.floor(1e4 * exclude_end[k][l]) / 1e4
                if s_val not in self.pp.exclude_start and e_val not in self.pp.exclude_end:
                    exclude_start2 = np.append(exclude_start2, math.floor(1e4 * s_val) / 1e4)
                    exclude_end2 = np.append(exclude_end2, math.floor(1e4 * e_val) / 1e4)
                    # self.pp.exclude_start = np.append(self.pp.exclude_start, math.floor(1e4 * s_val) / 1e4)
                    # self.pp.exclude_end = np.append(self.pp.exclude_end, math.floor(1e4 * e_val) / 1e4)

        # self.exclude_region()
        if self.pp.seg_align_ref_spc > 0:
            spc_idx = self.pp.seg_align_ref_spc - 1
        else:
            spc_idx = 0

        for k in range(len(exclude_start2)):
            idx = np.where((self.nmrdat[self.s][spc_idx].ppm1 > exclude_start2[k]) & (
                    self.nmrdat[self.s][spc_idx].ppm1 < exclude_end2[k]))
            self.deselect[idx] = np.ones(len(idx))

    # end segmental_alignment

    def select_plot_all(self):  # pragma: no cover
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].display.display_spc = True

        self.plot_spc()
        # end select_plot_all

    def select_plot_clear(self):  # pragma: no cover
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].display.display_spc = False

        self.plot_spc()
        # end select_plot_clear

    def set_autobaseline(self, autobaseline=False):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            self.nmrdat[self.s][k].proc.autobaseline = autobaseline

        return "set_autobaseline"

        # set_autobaseline

    def set_loadings_from_excel(self, file_name='', worksheet='', columns=[''], replace='', m0_factor=0.005, r2=0.0001):  # pragma: no cover
        if len(file_name) == 0:
            return

        if len(worksheet) == 0:
            return

        if len(columns) == 0:
            return

        if len(columns[0]) == 0:
            return

        try:
            xls = pd.ExcelFile(file_name)
        except:
            print('Filename not found')
            return

        try:
            df = pd.read_excel(xls, worksheet)
        except:
            return

        self.nmrdat.append([])
        n_sets = len(self.nmrdat)
        old_exp = self.e
        for k in range(len(self.nmrdat)):
            for l in range(len(self.nmrdat[self.s])):
                self.nmrdat[k][l].display.display_spc = False

        self.e = 0
        s = n_sets - 1
        if len(replace) == 0:
            select = np.where(self.nmrdat[self.s][self.e].spc[0] != 0)[0]
        else:
            ppm1 = np.copy(self.nmrdat[self.s][self.e].ppm1)
            ppm2 = df[df.keys()[0]]
            select = np.zeros(len(ppm2), dtype=int)
            for k in range(len(ppm2)):
                mdiff = np.abs(ppm1 - float(ppm2[k].replace(replace, '')))
                select[k] = np.where(mdiff == np.min(mdiff))[0][0]

        for k in range(len(columns)):
            nd1 = nd.NmrData()
            self.nmrdat[s].append(nd1)
            self.nmrdat[s][k].dim = 1
            self.nmrdat[s][k].fid = np.copy(self.nmrdat[self.s][self.e].fid)
            self.nmrdat[s][k].acq = self.nmrdat[self.s][self.e].acq
            self.nmrdat[s][k].display = self.nmrdat[self.s][self.e].display
            self.nmrdat[s][k].proc = self.nmrdat[self.s][self.e].proc
            self.nmrdat[s][k].ppm1 = np.resize(self.nmrdat[s][self.e].ppm1, (len(self.nmrdat[self.s][self.e].ppm1)))
            self.nmrdat[s][k].ppm1 = np.copy(self.nmrdat[self.s][self.e].ppm1)
            self.nmrdat[s][k].spc = np.resize(self.nmrdat[s][self.e].spc,
                                              (1, len(self.nmrdat[self.s][self.e].spc[0])))
            self.nmrdat[s][k].ref_shift = copy(self.nmrdat[self.s][self.e].ref_shift)
            self.nmrdat[s][k].ref_point = self.nmrdat[self.s][self.e].ref_point
            self.nmrdat[s][k].ref_point[0] = len(self.nmrdat[s][k].ppm1) - \
            np.where(np.abs(self.nmrdat[s][k].ppm1) == np.min(np.abs(self.nmrdat[s][k].ppm1)))[0][0]
            self.nmrdat[s][k].spc[0][select] = np.copy(df[columns[k]])
            self.nmrdat[s][k].title = 'Loadings from ' + columns[k] + '\n'
            m0 = np.max(self.nmrdat[s][k].spc[0].real) * m0_factor
            self.nmrdat[s][k].add_tmsp(m0, r2)

        self.e = old_exp
        # end set_loadings_from_excel

    def set_fit_ph1(self, fit_ph1=True):
        self.cf.fit_ph1 = fit_ph1
        self.cf.save_config()
        for k in range(len(self.nmrdat)):
            for l in range(len(self.nmrdat[k])):
                self.nmrdat[k][l].cf.read_config()

    # end set_fit_ph1

    def set_gb(self, gb):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            for l in range(len(gb)):
                self.nmrdat[self.s][k].proc.gb[l] = gb[l]

        return "set_gb"

    # end set_gb

    def set_lb(self, lb):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            for l in range(len(lb)):
                self.nmrdat[self.s][k].proc.lb[l] = lb[l]

        return "set_lb"

    # end set_lb

    def set_ph_from_exp(self, exp=-1):
        if exp == -1:
            exp = self.e

        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            if k != exp:
                self.nmrdat[self.s][k].proc.ph0 = self.nmrdat[self.s][exp].proc.ph0
                self.nmrdat[self.s][k].proc.ph1 = self.nmrdat[self.s][exp].proc.ph1

        return "set_ph_from_exp"

    # end set_ph_from_exp

    def set_ph0(self, ph0):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            for l in range(len(ph0)):
                self.nmrdat[self.s][k].proc.ph0[l] = ph0[l]

        return "set_ph0"

    # end set_ph0

    def set_ph1(self, ph1):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            for l in range(len(ph1)):
                self.nmrdat[self.s][k].proc.ph1[l] = ph1[l]

        return "set_ph1"

    # end set_ph1

    def set_n_points(self, n_points):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            for l in range(len(n_points)):
                self.nmrdat[self.s][k].proc.n_points[l] = n_points[l]

        return "set_n_points"

    # end set_n_points

    def set_ref(self, ref_value='auto'):
        self.nmrdat[self.s][self.e].ref = ref_value
        # end set_ref

    def set_ref_all(self, ref_value='auto'):
        for k in range(len(self.nmrdat[self.s])):
            self.nmrdat[self.s][k].ref = ref_value
        # end set_ref_all

    def set_ssb(self, ssb):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            for l in range(len(ssb)):
                self.nmrdat[self.s][k].proc.ssb[l] = ssb[l]

        return "set_ssb"

    # end set_ssb

    def set_standard_plot_colours(self):
        # self.cf.read_config()
        std_pos_col1 = (self.cf.pos_col10, self.cf.pos_col11, self.cf.pos_col12)
        std_neg_col1 = (self.cf.neg_col10, self.cf.neg_col11, self.cf.neg_col12)
        std_pos_col2 = (self.cf.pos_col20, self.cf.pos_col21, self.cf.pos_col22)
        std_neg_col2 = (self.cf.neg_col20, self.cf.neg_col21, self.cf.neg_col22)
        for k in range(len(self.nmrdat)):
            for l in range(len(self.nmrdat[k])):
                if self.cf.mode == 'dark' or (self.cf.mode == 'system' and darkdetect.isDark()):
                    self.nmrdat[k][l].display.pos_col_rgb = std_pos_col2
                    self.nmrdat[k][l].display.neg_col_rgb = std_neg_col2
                else:
                    self.nmrdat[k][l].display.pos_col_rgb = std_pos_col1
                    self.nmrdat[k][l].display.neg_col_rgb = std_neg_col1

    def set_water_suppression(self, ws='None'):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            self.nmrdat[self.s][k].proc.water_suppression = self.nmrdat[self.s][k].proc.water_supp[ws]

        return "set_water_suppression"

    # end set_ater_suppression

    def set_window_type(self, wt):
        n_exp = len(self.nmrdat[self.s])
        wt2 = []
        for k in range(len(wt)):
            if isinstance(wt[0], int):
                wt2.append(wt[k])
            else:
                wt2.append(self.nmrdat[self.s][self.e].proc.window_functions[wt[k]])

        for k in range(n_exp):
            for l in range(len(wt2)):
                self.nmrdat[self.s][k].proc.window_type[l] = wt2[l]

        return "set_window_type"

    # end set_window_type

    def set_zero_fill(self, zf):
        n_exp = len(self.nmrdat[self.s])
        for k in range(n_exp):
            for l in range(len(zf)):
                self.nmrdat[self.s][k].proc.n_points[l] = zf[l]

        return "set_zero_fill"

    # end set_zero_fill

    def shift_ref(self):
        for k in range(len(self.nmrdat[self.s])):
            shift_delta = self.nmrdat[self.s][k].ref_point[0] - self.nmrdat[self.s][0].ref_point[0]
            self.nmrdat[self.s][k].spc[0] = np.roll(self.nmrdat[self.s][k].spc[0], shift_delta)
            self.nmrdat[self.s][k].ppm1 = copy(self.nmrdat[self.s][0].ppm1)
            self.nmrdat[self.s][k].ref_shift = copy(self.nmrdat[self.s][0].ref_shift)

    # end shift_ref

    def spline_correct(self):
        if len(self.nmrdat[self.s][self.e].spline_baseline.baseline_points) > 0:
            for k in range(len(self.nmrdat[self.s])):
                self.nmrdat[self.s][k].spline_baseline.baseline_points = self.nmrdat[self.s][
                    self.e].spline_baseline.baseline_points
                self.nmrdat[self.s][k].corr_spline_baseline()

    def variance_stabilisation(self):
        if self.pp.auto_scaling:
            self.var_stab_auto_scale()

        if self.pp.pareto_scaling:
            self.var_stab_pareto_scale()

        if self.pp.g_log_transform:
            self.var_stab_g_log_transform()

    # end variance_stabilisation

    def var_stab_auto_scale(self):
        npts = len(self.nmrdat[self.s][0].spc[0])
        n_spc = len(self.nmrdat[self.s])
        spcs = np.zeros((n_spc, npts))
        for k in range(n_spc):
            spcs[k] = self.nmrdat[self.s][k].spc[0].real

        spc_mean = np.mean(spcs, 0)
        for k in range(n_spc):
            spcs[k] -= spc_mean

        spc_var = np.var(spcs, 0)
        for k in range(n_spc):
            idx = np.where(spc_var != 0)
            spcs[k][idx] = spcs[k][idx] / np.sqrt(spc_var[idx])
            self.nmrdat[self.s][k].spc[0] = spcs[k]

    # end var_stab_auto_scale

    def var_stab_g_log_transform(self):
        n_spc = len(self.nmrdat[self.s])
        l_min = np.zeros(n_spc)
        l_max = np.zeros(n_spc)
        for k in range(n_spc):
            l_max[k] = np.max(self.nmrdat[self.s][k].spc[0].real)

        m_max = np.max(l_max)
        for k in range(n_spc):
            spc = np.copy(self.nmrdat[self.s][k].spc[0].real)
            spc /= m_max
            spc = spc - self.pp.var_y0 + np.sqrt((spc - self.pp.var_y0) ** 2 + self.pp.var_lambda)
            idx = np.where(spc <= 0)
            spc[idx] = 1e-100
            self.nmrdat[self.s][k].spc[0] = np.copy(np.log(spc))
            l_min[k] = np.min(self.nmrdat[self.s][k].spc[0].real)

        for k in range(n_spc):
            self.nmrdat[self.s][k].spc[0] -= np.min(l_min)

    # end var_stab_g_log_transform

    def var_stab_pareto_scale(self):
        npts = len(self.nmrdat[self.s][0].spc[0])
        n_spc = len(self.nmrdat[self.s])
        spcs = np.zeros((n_spc, npts))
        for k in range(n_spc):
            spcs[k] = self.nmrdat[self.s][k].spc[0].real

        spc_var = np.sqrt(np.std(spcs, 0))
        for k in range(n_spc):
            idx = np.where(spc_var != 0)
            spcs[k][idx] = spcs[k][idx] / spc_var[idx]
            self.nmrdat[self.s][k].spc[0] = spcs[k]

    # end var_stab_pareto_scale
